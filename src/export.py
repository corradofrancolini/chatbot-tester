"""
Export Module - Generazione Report

Supporta:
- PDF: report professionale con grafici e screenshot
- Excel: dati strutturati per analisi
- HTML: file standalone (esistente, migliorato)

Dipendenze opzionali:
- PDF: reportlab, pillow
- Excel: openpyxl
"""

from dataclasses import dataclass, field
from typing import Optional, List, Dict, Any, Union
from pathlib import Path
from datetime import datetime
import json
import csv
import base64
import io

# Import opzionali per PDF
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, letter
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch, cm
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
        Image, PageBreak, KeepTogether
    )
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

# Import opzionali per Excel
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference, PieChart
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# Import opzionali per immagini
try:
    from PIL import Image as PILImage
    PIL_AVAILABLE = True
except ImportError:
    PIL_AVAILABLE = False


@dataclass
class TestResult:
    """Risultato singolo test per export"""
    test_id: str
    category: str
    question: str
    expected: str
    actual_response: str
    status: str  # PASS, FAIL, SKIP
    score: Optional[float] = None
    evaluation: Optional[str] = None
    sources: List[str] = field(default_factory=list)
    screenshot_path: Optional[str] = None
    duration_seconds: float = 0
    conversation: List[Dict[str, str]] = field(default_factory=list)


@dataclass
class RunReport:
    """Report completo di un run"""
    project: str
    run_number: int
    timestamp: str
    env: str
    prompt_version: str
    model_version: str
    total_tests: int
    passed: int
    failed: int
    skipped: int
    pass_rate: float
    duration_seconds: float
    tests: List[TestResult] = field(default_factory=list)
    regressions: List[str] = field(default_factory=list)

    @classmethod
    def from_local_report(cls, report_path: Path) -> 'RunReport':
        """Carica da report locale JSON"""
        with open(report_path) as f:
            data = json.load(f)

        tests = []
        for t in data.get('tests', []):
            tests.append(TestResult(
                test_id=t.get('test_id', ''),
                category=t.get('category', ''),
                question=t.get('question', ''),
                expected=t.get('expected', ''),
                actual_response=t.get('response', ''),
                status=t.get('status', 'SKIP'),
                score=t.get('score'),
                evaluation=t.get('evaluation'),
                sources=t.get('sources', []),
                screenshot_path=t.get('screenshot'),
                duration_seconds=t.get('duration', 0),
                conversation=t.get('conversation', [])
            ))

        return cls(
            project=data.get('project', ''),
            run_number=data.get('run_number', 0),
            timestamp=data.get('timestamp', ''),
            env=data.get('env', ''),
            prompt_version=data.get('prompt_version', ''),
            model_version=data.get('model_version', ''),
            total_tests=data.get('total_tests', 0),
            passed=data.get('passed', 0),
            failed=data.get('failed', 0),
            skipped=data.get('skipped', 0),
            pass_rate=data.get('pass_rate', 0),
            duration_seconds=data.get('duration', 0),
            tests=tests,
            regressions=data.get('regressions', [])
        )

    @classmethod
    def from_summary_and_csv(cls, summary_path: Path, csv_path: Path = None) -> 'RunReport':
        """Carica da summary.json e report.csv (formato esistente)"""
        with open(summary_path) as f:
            summary = json.load(f)

        tests = []
        screenshots_dir = summary_path.parent / "screenshots"

        # Carica test da CSV se disponibile
        if csv_path and csv_path.exists():
            import csv as csv_module
            with open(csv_path, 'r', encoding='utf-8') as f:
                reader = csv_module.DictReader(f)
                for row in reader:
                    test_id = row.get('test_id', row.get('TEST_ID', ''))
                    screenshot_path = None
                    if screenshots_dir.exists():
                        potential_screenshot = screenshots_dir / f"{test_id}.png"
                        if potential_screenshot.exists():
                            screenshot_path = str(potential_screenshot)

                    tests.append(TestResult(
                        test_id=test_id,
                        category=row.get('category', row.get('CATEGORY', 'uncategorized')),
                        question=row.get('question', row.get('QUESTION', '')),
                        expected=row.get('expected', row.get('EXPECTED', '')),
                        actual_response=row.get('response', row.get('RESPONSE', '')),
                        status=row.get('status', row.get('STATUS', 'SKIP')),
                        score=float(row.get('score', 0)) if row.get('score') else None,
                        evaluation=row.get('evaluation', row.get('EVALUATION', '')),
                        sources=[],
                        screenshot_path=screenshot_path,
                        duration_seconds=float(row.get('duration', 0)) if row.get('duration') else 0,
                        conversation=[]
                    ))

        total = summary.get('total_tests', len(tests))
        passed = summary.get('passed', 0)
        pass_rate = (passed / total * 100) if total > 0 else 0

        return cls(
            project=summary.get('project_name', ''),
            run_number=summary.get('run_number', 0),
            timestamp=summary.get('start_time', ''),
            env=summary.get('env', 'DEV'),
            prompt_version=summary.get('prompt_version', ''),
            model_version=summary.get('model_version', ''),
            total_tests=total,
            passed=passed,
            failed=summary.get('failed', 0),
            skipped=summary.get('skipped', 0),
            pass_rate=pass_rate,
            duration_seconds=summary.get('duration_seconds', 0),
            tests=tests,
            regressions=[]
        )


class PDFExporter:
    """Esportazione report in PDF"""

    def __init__(self, output_path: Path):
        if not REPORTLAB_AVAILABLE:
            raise ImportError("reportlab non installato. Installa con: pip install reportlab")
        self.output_path = output_path
        self.styles = getSampleStyleSheet()
        self._setup_custom_styles()

    def _setup_custom_styles(self):
        """Configura stili personalizzati"""
        self.styles.add(ParagraphStyle(
            name='Title',
            parent=self.styles['Heading1'],
            fontSize=24,
            spaceAfter=30,
            alignment=TA_CENTER
        ))
        self.styles.add(ParagraphStyle(
            name='Subtitle',
            parent=self.styles['Normal'],
            fontSize=12,
            textColor=colors.grey,
            alignment=TA_CENTER,
            spaceAfter=20
        ))
        self.styles.add(ParagraphStyle(
            name='SectionHeader',
            parent=self.styles['Heading2'],
            fontSize=14,
            spaceBefore=20,
            spaceAfter=10,
            textColor=colors.HexColor('#2c3e50')
        ))
        self.styles.add(ParagraphStyle(
            name='TestQuestion',
            parent=self.styles['Normal'],
            fontSize=10,
            leftIndent=20,
            textColor=colors.HexColor('#34495e')
        ))

    def export(self, report: RunReport, include_screenshots: bool = True) -> Path:
        """Genera PDF del report"""
        doc = SimpleDocTemplate(
            str(self.output_path),
            pagesize=A4,
            rightMargin=1.5*cm,
            leftMargin=1.5*cm,
            topMargin=2*cm,
            bottomMargin=2*cm
        )

        story = []

        # Titolo
        story.append(Paragraph(
            f"Chatbot Test Report",
            self.styles['Title']
        ))
        story.append(Paragraph(
            f"{report.project} - RUN {report.run_number}",
            self.styles['Subtitle']
        ))
        story.append(Spacer(1, 20))

        # Executive Summary
        story.append(Paragraph("Executive Summary", self.styles['SectionHeader']))
        story.extend(self._create_summary_section(report))
        story.append(Spacer(1, 20))

        # Tabella risultati
        story.append(Paragraph("Risultati per Test", self.styles['SectionHeader']))
        story.extend(self._create_results_table(report))

        # Regressioni (se presenti)
        if report.regressions:
            story.append(PageBreak())
            story.append(Paragraph("Regressioni Rilevate", self.styles['SectionHeader']))
            story.extend(self._create_regressions_section(report))

        # Dettaglio test falliti
        failed_tests = [t for t in report.tests if t.status == 'FAIL']
        if failed_tests:
            story.append(PageBreak())
            story.append(Paragraph("Dettaglio Test Falliti", self.styles['SectionHeader']))
            story.extend(self._create_failed_tests_section(failed_tests, include_screenshots))

        # Footer info
        story.append(Spacer(1, 30))
        story.append(Paragraph(
            f"Generato il {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} da Chatbot Tester",
            ParagraphStyle('Footer', parent=self.styles['Normal'], fontSize=8, textColor=colors.grey)
        ))

        doc.build(story)
        return self.output_path

    def _create_summary_section(self, report: RunReport) -> List:
        """Crea sezione summary"""
        elements = []

        # Info generali
        info_data = [
            ['Progetto', report.project],
            ['Run', f"#{report.run_number}"],
            ['Data', report.timestamp[:10] if report.timestamp else 'N/A'],
            ['Ambiente', report.env],
            ['Prompt Version', report.prompt_version],
            ['Model Version', report.model_version],
        ]

        info_table = Table(info_data, colWidths=[3*cm, 6*cm])
        info_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        elements.append(info_table)
        elements.append(Spacer(1, 15))

        # Metriche
        pass_color = colors.HexColor('#28a745') if report.pass_rate >= 80 else colors.HexColor('#dc3545')

        metrics_data = [
            ['Totale Test', 'Passati', 'Falliti', 'Pass Rate'],
            [str(report.total_tests), str(report.passed), str(report.failed), f"{report.pass_rate:.1f}%"]
        ]

        metrics_table = Table(metrics_data, colWidths=[3.5*cm, 3.5*cm, 3.5*cm, 3.5*cm])
        metrics_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f8f9fa')),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 11),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#dee2e6')),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
            ('TOPPADDING', (0, 0), (-1, -1), 10),
            ('TEXTCOLOR', (3, 1), (3, 1), pass_color),
            ('FONTNAME', (3, 1), (3, 1), 'Helvetica-Bold'),
        ]))
        elements.append(metrics_table)

        return elements

    def _create_results_table(self, report: RunReport) -> List:
        """Crea tabella risultati"""
        elements = []

        # Header
        data = [['ID', 'Categoria', 'Status', 'Score']]

        # Righe
        for test in report.tests:
            status_text = test.status
            score_text = f"{test.score:.1f}" if test.score is not None else "-"
            data.append([test.test_id, test.category[:20], status_text, score_text])

        table = Table(data, colWidths=[2.5*cm, 6*cm, 2.5*cm, 2*cm])

        # Stile con colori per status
        style = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#343a40')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('ALIGN', (2, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#dee2e6')),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
        ]

        # Colora righe per status
        for i, test in enumerate(report.tests, start=1):
            if test.status == 'PASS':
                style.append(('TEXTCOLOR', (2, i), (2, i), colors.HexColor('#28a745')))
            elif test.status == 'FAIL':
                style.append(('TEXTCOLOR', (2, i), (2, i), colors.HexColor('#dc3545')))
                style.append(('BACKGROUND', (0, i), (-1, i), colors.HexColor('#fff5f5')))

        table.setStyle(TableStyle(style))
        elements.append(table)

        return elements

    def _create_regressions_section(self, report: RunReport) -> List:
        """Crea sezione regressioni"""
        elements = []

        for regression in report.regressions:
            elements.append(Paragraph(
                f"- {regression}",
                self.styles['TestQuestion']
            ))

        return elements

    def _create_failed_tests_section(self, tests: List[TestResult], include_screenshots: bool) -> List:
        """Crea sezione dettaglio test falliti"""
        elements = []

        for test in tests:
            # Box per ogni test
            test_elements = []

            test_elements.append(Paragraph(
                f"<b>{test.test_id}</b> - {test.category}",
                self.styles['Heading3']
            ))

            test_elements.append(Paragraph(
                f"<b>Domanda:</b> {test.question[:500]}",
                self.styles['Normal']
            ))

            test_elements.append(Paragraph(
                f"<b>Risposta attesa:</b> {test.expected[:300]}",
                self.styles['Normal']
            ))

            test_elements.append(Paragraph(
                f"<b>Risposta ricevuta:</b> {test.actual_response[:500]}",
                self.styles['Normal']
            ))

            if test.evaluation:
                test_elements.append(Paragraph(
                    f"<b>Valutazione:</b> {test.evaluation[:300]}",
                    self.styles['Normal']
                ))

            # Screenshot (se disponibile e richiesto)
            if include_screenshots and test.screenshot_path and PIL_AVAILABLE:
                screenshot_path = Path(test.screenshot_path)
                if screenshot_path.exists():
                    try:
                        img = Image(str(screenshot_path), width=12*cm, height=8*cm)
                        test_elements.append(Spacer(1, 10))
                        test_elements.append(img)
                    except Exception:
                        pass

            test_elements.append(Spacer(1, 20))

            # Mantieni insieme
            elements.append(KeepTogether(test_elements))

        return elements


class ExcelExporter:
    """Esportazione report in Excel"""

    def __init__(self, output_path: Path):
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl non installato. Installa con: pip install openpyxl")
        self.output_path = output_path

    def export(self, report: RunReport) -> Path:
        """Genera Excel del report"""
        wb = Workbook()

        # Sheet Summary
        self._create_summary_sheet(wb, report)

        # Sheet Risultati
        self._create_results_sheet(wb, report)

        # Sheet Dettaglio (conversazioni)
        self._create_detail_sheet(wb, report)

        # Rimuovi sheet default
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']

        wb.save(str(self.output_path))
        return self.output_path

    def _create_summary_sheet(self, wb: Workbook, report: RunReport):
        """Crea sheet summary"""
        ws = wb.create_sheet("Summary", 0)

        # Stili
        header_font = Font(bold=True, size=14)
        label_font = Font(bold=True)
        pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        # Titolo
        ws['A1'] = f"Chatbot Test Report - {report.project}"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:D1')

        # Info generali
        info = [
            ('Run', report.run_number),
            ('Data', report.timestamp),
            ('Ambiente', report.env),
            ('Prompt Version', report.prompt_version),
            ('Model Version', report.model_version),
        ]

        row = 3
        for label, value in info:
            ws.cell(row=row, column=1, value=label).font = label_font
            ws.cell(row=row, column=2, value=value)
            row += 1

        # Metriche
        row += 1
        ws.cell(row=row, column=1, value="Metriche").font = header_font
        row += 1

        metrics = [
            ('Totale Test', report.total_tests),
            ('Passati', report.passed),
            ('Falliti', report.failed),
            ('Skipped', report.skipped),
            ('Pass Rate', f"{report.pass_rate:.1f}%"),
            ('Durata (s)', report.duration_seconds),
        ]

        for label, value in metrics:
            ws.cell(row=row, column=1, value=label).font = label_font
            cell = ws.cell(row=row, column=2, value=value)
            if label == 'Passati':
                cell.fill = pass_fill
            elif label == 'Falliti' and report.failed > 0:
                cell.fill = fail_fill
            row += 1

        # Regressioni
        if report.regressions:
            row += 1
            ws.cell(row=row, column=1, value="Regressioni").font = header_font
            row += 1
            for reg in report.regressions:
                ws.cell(row=row, column=1, value=reg)
                row += 1

        # Adatta colonne
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30

    def _create_results_sheet(self, wb: Workbook, report: RunReport):
        """Crea sheet risultati"""
        ws = wb.create_sheet("Risultati")

        # Header
        headers = ['Test ID', 'Categoria', 'Status', 'Score', 'Domanda', 'Risposta Attesa', 'Risposta Ricevuta']
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="343A40", end_color="343A40", fill_type="solid")

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        # Dati
        pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        for row, test in enumerate(report.tests, start=2):
            ws.cell(row=row, column=1, value=test.test_id)
            ws.cell(row=row, column=2, value=test.category)

            status_cell = ws.cell(row=row, column=3, value=test.status)
            if test.status == 'PASS':
                status_cell.fill = pass_fill
            elif test.status == 'FAIL':
                status_cell.fill = fail_fill

            ws.cell(row=row, column=4, value=test.score if test.score else '')
            ws.cell(row=row, column=5, value=test.question[:500])
            ws.cell(row=row, column=6, value=test.expected[:500])
            ws.cell(row=row, column=7, value=test.actual_response[:500])

        # Adatta colonne
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 8
        ws.column_dimensions['E'].width = 40
        ws.column_dimensions['F'].width = 40
        ws.column_dimensions['G'].width = 40

        # Filtri
        ws.auto_filter.ref = ws.dimensions

    def _create_detail_sheet(self, wb: Workbook, report: RunReport):
        """Crea sheet dettaglio conversazioni"""
        ws = wb.create_sheet("Dettaglio")

        headers = ['Test ID', 'Turno', 'Ruolo', 'Messaggio', 'Sources']
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        row = 2
        for test in report.tests:
            if not test.conversation:
                continue

            for turn_idx, turn in enumerate(test.conversation, start=1):
                ws.cell(row=row, column=1, value=test.test_id)
                ws.cell(row=row, column=2, value=turn_idx)
                ws.cell(row=row, column=3, value=turn.get('role', ''))
                ws.cell(row=row, column=4, value=turn.get('content', '')[:1000])
                ws.cell(row=row, column=5, value=', '.join(test.sources) if turn_idx == 1 else '')
                row += 1

        # Adatta colonne
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 8
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 80
        ws.column_dimensions['E'].width = 40


class HTMLExporter:
    """Esportazione report in HTML standalone"""

    def __init__(self, output_path: Path):
        self.output_path = output_path

    def export(self, report: RunReport, include_screenshots: bool = True) -> Path:
        """Genera HTML del report"""
        html = self._generate_html(report, include_screenshots)

        with open(self.output_path, 'w', encoding='utf-8') as f:
            f.write(html)

        return self.output_path

    def _generate_html(self, report: RunReport, include_screenshots: bool) -> str:
        """Genera contenuto HTML"""
        # Embed screenshots come base64
        screenshots_html = ""
        if include_screenshots:
            for test in report.tests:
                if test.screenshot_path and Path(test.screenshot_path).exists():
                    try:
                        with open(test.screenshot_path, 'rb') as f:
                            b64 = base64.b64encode(f.read()).decode()
                            test.screenshot_base64 = f"data:image/png;base64,{b64}"
                    except Exception:
                        test.screenshot_base64 = None
                else:
                    test.screenshot_base64 = None

        pass_color = "#28a745" if report.pass_rate >= 80 else "#dc3545"

        tests_html = ""
        for test in report.tests:
            status_class = "pass" if test.status == "PASS" else "fail" if test.status == "FAIL" else "skip"
            screenshot_img = ""
            if include_screenshots and hasattr(test, 'screenshot_base64') and test.screenshot_base64:
                screenshot_img = f'<img src="{test.screenshot_base64}" style="max-width: 100%; margin-top: 10px; border: 1px solid #ddd;">'

            tests_html += f"""
            <div class="test-card {status_class}">
                <div class="test-header">
                    <span class="test-id">{test.test_id}</span>
                    <span class="test-status">{test.status}</span>
                </div>
                <div class="test-category">{test.category}</div>
                <div class="test-content">
                    <p><strong>Domanda:</strong> {test.question[:500]}</p>
                    <p><strong>Risposta:</strong> {test.actual_response[:500]}</p>
                    {f'<p><strong>Valutazione:</strong> {test.evaluation}</p>' if test.evaluation else ''}
                    {screenshot_img}
                </div>
            </div>
            """

        html = f"""
<!DOCTYPE html>
<html lang="it">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Test Report - {report.project} RUN {report.run_number}</title>
    <style>
        * {{ box-sizing: border-box; margin: 0; padding: 0; }}
        body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif; background: #f5f5f5; padding: 20px; }}
        .container {{ max-width: 1200px; margin: 0 auto; }}
        .header {{ background: linear-gradient(135deg, #2c3e50, #34495e); color: white; padding: 30px; border-radius: 10px; margin-bottom: 20px; }}
        .header h1 {{ font-size: 28px; margin-bottom: 10px; }}
        .header .subtitle {{ opacity: 0.8; }}
        .summary {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 20px; margin-bottom: 20px; }}
        .metric {{ background: white; padding: 20px; border-radius: 10px; text-align: center; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }}
        .metric .value {{ font-size: 36px; font-weight: bold; color: #2c3e50; }}
        .metric .label {{ color: #7f8c8d; margin-top: 5px; }}
        .metric.pass-rate .value {{ color: {pass_color}; }}
        .info-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(250px, 1fr)); gap: 15px; background: white; padding: 20px; border-radius: 10px; margin-bottom: 20px; }}
        .info-item {{ display: flex; justify-content: space-between; padding: 8px 0; border-bottom: 1px solid #eee; }}
        .info-item:last-child {{ border-bottom: none; }}
        .tests-section {{ margin-top: 30px; }}
        .tests-section h2 {{ margin-bottom: 20px; color: #2c3e50; }}
        .test-card {{ background: white; border-radius: 10px; margin-bottom: 15px; overflow: hidden; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }}
        .test-card.pass {{ border-left: 4px solid #28a745; }}
        .test-card.fail {{ border-left: 4px solid #dc3545; }}
        .test-card.skip {{ border-left: 4px solid #ffc107; }}
        .test-header {{ display: flex; justify-content: space-between; align-items: center; padding: 15px 20px; background: #f8f9fa; }}
        .test-id {{ font-weight: bold; color: #2c3e50; }}
        .test-status {{ padding: 4px 12px; border-radius: 20px; font-size: 12px; font-weight: bold; }}
        .pass .test-status {{ background: #d4edda; color: #155724; }}
        .fail .test-status {{ background: #f8d7da; color: #721c24; }}
        .test-category {{ padding: 5px 20px; font-size: 12px; color: #6c757d; background: #f8f9fa; }}
        .test-content {{ padding: 20px; }}
        .test-content p {{ margin-bottom: 10px; line-height: 1.6; }}
        .footer {{ text-align: center; padding: 20px; color: #7f8c8d; font-size: 12px; }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>Chatbot Test Report</h1>
            <div class="subtitle">{report.project} - RUN {report.run_number}</div>
        </div>

        <div class="summary">
            <div class="metric">
                <div class="value">{report.total_tests}</div>
                <div class="label">Totale Test</div>
            </div>
            <div class="metric">
                <div class="value" style="color: #28a745;">{report.passed}</div>
                <div class="label">Passati</div>
            </div>
            <div class="metric">
                <div class="value" style="color: #dc3545;">{report.failed}</div>
                <div class="label">Falliti</div>
            </div>
            <div class="metric pass-rate">
                <div class="value">{report.pass_rate:.1f}%</div>
                <div class="label">Pass Rate</div>
            </div>
        </div>

        <div class="info-grid">
            <div class="info-item"><span>Data</span><span>{report.timestamp[:10] if report.timestamp else 'N/A'}</span></div>
            <div class="info-item"><span>Ambiente</span><span>{report.env}</span></div>
            <div class="info-item"><span>Prompt Version</span><span>{report.prompt_version}</span></div>
            <div class="info-item"><span>Model Version</span><span>{report.model_version}</span></div>
            <div class="info-item"><span>Durata</span><span>{report.duration_seconds:.0f}s</span></div>
        </div>

        <div class="tests-section">
            <h2>Risultati Test</h2>
            {tests_html}
        </div>

        <div class="footer">
            Generato il {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} da Chatbot Tester
        </div>
    </div>
</body>
</html>
        """

        return html


class ReportExporter:
    """
    Gestore centralizzato export report.

    Usage:
        exporter = ReportExporter(report)

        # Export singolo formato
        exporter.to_pdf(Path("report.pdf"))
        exporter.to_excel(Path("report.xlsx"))
        exporter.to_html(Path("report.html"))

        # Export multiplo
        exporter.export_all(output_dir)
    """

    def __init__(self, report: RunReport):
        self.report = report

    def to_pdf(self, output_path: Path, include_screenshots: bool = True) -> Path:
        """Esporta in PDF"""
        exporter = PDFExporter(output_path)
        return exporter.export(self.report, include_screenshots)

    def to_excel(self, output_path: Path) -> Path:
        """Esporta in Excel"""
        exporter = ExcelExporter(output_path)
        return exporter.export(self.report)

    def to_html(self, output_path: Path, include_screenshots: bool = True) -> Path:
        """Esporta in HTML"""
        exporter = HTMLExporter(output_path)
        return exporter.export(self.report, include_screenshots)

    def to_csv(self, output_path: Path) -> Path:
        """Esporta in CSV (semplice)"""
        with open(output_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(['Test ID', 'Categoria', 'Status', 'Score', 'Domanda', 'Risposta'])

            for test in self.report.tests:
                writer.writerow([
                    test.test_id,
                    test.category,
                    test.status,
                    test.score or '',
                    test.question[:200],
                    test.actual_response[:200]
                ])

        return output_path

    def export_all(self, output_dir: Path) -> Dict[str, Path]:
        """Esporta in tutti i formati disponibili"""
        output_dir.mkdir(parents=True, exist_ok=True)

        base_name = f"{self.report.project}_run{self.report.run_number}"
        results = {}

        # HTML (sempre disponibile)
        results['html'] = self.to_html(output_dir / f"{base_name}.html")

        # CSV (sempre disponibile)
        results['csv'] = self.to_csv(output_dir / f"{base_name}.csv")

        # PDF (se reportlab disponibile)
        if REPORTLAB_AVAILABLE:
            try:
                results['pdf'] = self.to_pdf(output_dir / f"{base_name}.pdf")
            except Exception as e:
                print(f"! PDF export failed: {e}")

        # Excel (se openpyxl disponibile)
        if OPENPYXL_AVAILABLE:
            try:
                results['excel'] = self.to_excel(output_dir / f"{base_name}.xlsx")
            except Exception as e:
                print(f"! Excel export failed: {e}")

        return results


def check_dependencies() -> Dict[str, bool]:
    """Verifica dipendenze disponibili"""
    return {
        'pdf': REPORTLAB_AVAILABLE,
        'excel': OPENPYXL_AVAILABLE,
        'images': PIL_AVAILABLE
    }


if __name__ == "__main__":
    # Test
    deps = check_dependencies()
    print("Dipendenze disponibili:")
    for name, available in deps.items():
        status = "OK" if available else "Non installato"
        print(f"  {name}: {status}")

    if not deps['pdf']:
        print("\nPer PDF: pip install reportlab pillow")
    if not deps['excel']:
        print("Per Excel: pip install openpyxl")
