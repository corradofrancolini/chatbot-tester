"""
MCP Tools for chatbot-tester.

Defines all tools exposed by the MCP server for interacting with
chatbot-tester functionality.
"""

import os
import sys
import json
import logging
from pathlib import Path
from typing import Optional
from datetime import datetime

# Add parent directory to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent))

from mcp.server import Server
from mcp.types import Tool, TextContent

logger = logging.getLogger(__name__)


def get_field(record: dict, field: str, default: str = "") -> str:
    """
    Get a field from a record dictionary, case-insensitive.

    Google Sheets headers may be uppercase (ESITO) but code may use lowercase (esito).
    Headers may also have spaces (TEST ID) vs underscores (test_id).
    This helper tries multiple variations.
    """
    # Try exact match first
    if field in record:
        return record[field]

    # Try case variations
    variations = [
        field.upper(),           # ESITO
        field.lower(),           # esito
        field.title(),           # Esito
        field.replace("_", " "), # test id
        field.replace("_", " ").upper(),    # TEST ID
        field.replace("_", " ").title(),    # Test Id
        field.replace(" ", "_"),            # test_id
        field.replace(" ", "_").upper(),    # TEST_ID
    ]

    for variant in variations:
        if variant in record:
            return record[variant]

    return default


# Import version from package
from mcp_server import __version__

# Projects directory
PROJECTS_DIR = Path(__file__).parent.parent / "projects"

# Changelog MCP - scritto per l'utente finale (il collega)
# La data della versione corrente viene generata automaticamente
TODAY = datetime.now().strftime("%d/%m/%Y")

CHANGELOG = [
    {
        "version": __version__,
        "date": TODAY,
        "title": "Analytics e Report avanzati",
        "changes": [
            {
                "icon": "📊",
                "title": "Analisi Flaky e Regressioni",
                "description": "Rileva test instabili con 'Test flaky?' e regressioni con 'Cosa si è rotto?'. Confronta automaticamente le ultime run."
            },
            {
                "icon": "⚡",
                "title": "Performance e Alert",
                "description": "Monitora timing, throughput e latenze con 'Performance report'. Ricevi alert automatici su soglie superate."
            },
            {
                "icon": "📈",
                "title": "Coverage e Stabilità",
                "description": "Analizza gap di copertura per categoria e genera report di stabilità della suite di test."
            },
            {
                "icon": "🔧",
                "title": "Diagnosi e Calibrazione",
                "description": "Diagnosi intelligente dei fallimenti con 'Perché fallisce?' e calibrazione automatica delle soglie."
            },
            {
                "icon": "📄",
                "title": "Export Report",
                "description": "Esporta i risultati in Excel, HTML o CSV con 'Esporta in Excel'."
            },
        ]
    },
    {
        "version": "1.3.0",
        "date": "22/12/2024",
        "title": "Comunicazione e gestione test",
        "changes": [
            {
                "icon": "🔔",
                "title": "Contatto diretto",
                "description": "Puoi contattare direttamente Corrado via Telegram con 'Avvisa Corrado che...' per segnalare problemi, fare domande o richiedere funzionalità."
            },
            {
                "icon": "📝",
                "title": "Creazione test",
                "description": "Puoi aggiungere nuovi test case direttamente dalla chat con 'Aggiungi questo test a silicon-b'. Specifica domanda, risposta attesa e categoria."
            },
            {
                "icon": "📋",
                "title": "Lista completa",
                "description": "La lista test ora mostra tutti i test disponibili, non più limitata ai primi 10."
            },
        ]
    },
    {
        "version": "1.2.0",
        "date": "21/12/2024",
        "title": "Sessioni guidate",
        "changes": [
            {
                "icon": "🚀",
                "title": "Wizard interattivo",
                "description": "Scrivi 'Voglio testare silicon-b' per avviare una sessione guidata che ti mostra test set disponibili e stato attuale."
            },
            {
                "icon": "📊",
                "title": "Conferma prima dell'esecuzione",
                "description": "Prima di lanciare i test vedi un riepilogo dettagliato con numero test, tempo stimato e configurazione."
            },
            {
                "icon": "⚡",
                "title": "Parallelismo automatico",
                "description": "Per test set con più di 10 test, l'esecuzione usa automaticamente 3 container paralleli per velocizzare."
            },
        ]
    },
    {
        "version": "1.1.0",
        "date": "20/12/2024",
        "title": "Prima release",
        "changes": [
            {
                "icon": "✨",
                "title": "Funzionalità base",
                "description": "Lancia test su CircleCI, controlla lo stato delle pipeline, visualizza i risultati da Google Sheets."
            },
        ]
    },
]


def get_available_projects() -> list[str]:
    """Get list of available project names."""
    if not PROJECTS_DIR.exists():
        return []

    projects = []
    for item in PROJECTS_DIR.iterdir():
        if item.is_dir() and not item.name.startswith('.'):
            # Check if it has a project.yaml
            if (item / "project.yaml").exists():
                projects.append(item.name)

    return sorted(projects)


def get_available_test_sets(project: str) -> dict[str, int]:
    """Get available test sets for a project with their test counts."""
    project_dir = PROJECTS_DIR / project
    test_sets = {}

    if not project_dir.exists():
        return test_sets

    # Look for test files
    for file in project_dir.iterdir():
        if file.name.startswith("tests") and file.name.endswith(".json"):
            try:
                with open(file) as f:
                    data = json.load(f)
                    tests = data if isinstance(data, list) else data.get("tests", [])
                    # Create friendly name
                    if file.name == "tests.json":
                        name = "standard"
                    else:
                        # tests_paraphrase.json -> paraphrase
                        name = file.name.replace("tests_", "").replace(".json", "")
                    test_sets[name] = len(tests)
            except Exception as e:
                logger.error(f"Error reading {file}: {e}")

    return test_sets


def get_project_tests(project: str, test_set: str = "standard") -> list[dict]:
    """Get tests for a project from a specific test set."""
    project_dir = PROJECTS_DIR / project

    # Map test set name to file
    if test_set == "standard":
        tests_file = project_dir / "tests.json"
    else:
        tests_file = project_dir / f"tests_{test_set}.json"

    if not tests_file.exists():
        # Fallback to standard tests.json
        tests_file = project_dir / "tests.json"
        if not tests_file.exists():
            return []

    try:
        with open(tests_file) as f:
            data = json.load(f)
            # Handle both formats: direct array [...] or object {"tests": [...]}
            if isinstance(data, list):
                return data
            return data.get("tests", [])
    except Exception as e:
        logger.error(f"Error reading tests: {e}")
        return []


def get_run_config(project: str) -> dict:
    """Get run_config.json for a project."""
    config_file = PROJECTS_DIR / project / "run_config.json"
    if not config_file.exists():
        return {}
    try:
        with open(config_file) as f:
            return json.load(f)
    except Exception as e:
        logger.error(f"Error reading run_config: {e}")
        return {}


def should_use_parallel(test_count: int) -> bool:
    """Decide automaticamente se usare parallelismo."""
    return test_count >= 10


def estimate_duration(test_count: int, parallel: bool) -> str:
    """Stima durata esecuzione."""
    # ~1.5 min per test singolo, ~0.5 min con 3 container
    if parallel:
        minutes = (test_count / 3) * 0.5 + 2  # +2 min overhead
    else:
        minutes = test_count * 1.5

    if minutes < 5:
        return "5 minuti"
    elif minutes < 15:
        return "10-15 minuti"
    else:
        return f"{int(minutes)} minuti"


def register_tools(server: Server):
    """Register all MCP tools with the server."""

    @server.list_tools()
    async def list_tools() -> list[Tool]:
        """List available tools."""
        tools = [
            Tool(
                name="get_help",
                description="""Mostra la guida su come usare chatbot-tester.

USA QUESTO TOOL quando l'utente chiede:
- "Come funziona?"
- "Aiuto"
- "Cosa posso fare?"
- "Come si usa?"

Se l'utente è nuovo, usa topic='quickstart' per una guida rapida.""",
                inputSchema={
                    "type": "object",
                    "properties": {
                        "topic": {
                            "type": "string",
                            "enum": ["overview", "testing", "results", "comparison", "quickstart", "all"],
                            "description": "Argomento specifico (default: 'all'). Usa 'quickstart' per utenti nuovi.",
                            "default": "all"
                        }
                    },
                    "required": []
                }
            ),
            Tool(
                name="list_projects",
                description="""Mostra i progetti disponibili per il testing.

USA QUESTO TOOL quando l'utente chiede:
- "Quali progetti posso testare?"
- "Cosa c'è disponibile?"
- "Lista progetti"
- "Mostrami i progetti"

Per ogni progetto mostra i test set disponibili (es. standard, paraphrase, ggp).""",
                inputSchema={
                    "type": "object",
                    "properties": {},
                    "required": []
                }
            ),
            Tool(
                name="suggest_project",
                description="Analizza i progetti disponibili e suggerisce quale testare, mostrando statistiche e priorità",
                inputSchema={
                    "type": "object",
                    "properties": {},
                    "required": []
                }
            ),
            Tool(
                name="list_test_sets",
                description="Mostra i set di test disponibili per un progetto (es. standard, paraphrase)",
                inputSchema={
                    "type": "object",
                    "properties": {
                        "project": {
                            "type": "string",
                            "description": "Nome del progetto (es. 'silicon-b')"
                        }
                    },
                    "required": ["project"]
                }
            ),
            Tool(
                name="list_tests",
                description="Elenca i test disponibili per un progetto, con filtro opzionale per ID o keyword",
                inputSchema={
                    "type": "object",
                    "properties": {
                        "project": {
                            "type": "string",
                            "description": "Nome del progetto (es. 'silicon-b')"
                        },
                        "test_set": {
                            "type": "string",
                            "description": "Nome del set di test (corrisponde al file tests_{name}.json). Default: 'standard'"
                        },
                        "filter": {
                            "type": "string",
                            "description": "Filtra test per ID o keyword (es. 'PARA', 'TEST_001', 'prodotto')"
                        }
                    },
                    "required": ["project"]
                }
            ),
            Tool(
                name="trigger_circleci",
                description="Avvia una pipeline CircleCI per eseguire i test su un progetto. "
                           "Questo esegue i test nel cloud e scrive i risultati su Google Sheets.",
                inputSchema={
                    "type": "object",
                    "properties": {
                        "project": {
                            "type": "string",
                            "description": "Nome del progetto (es. 'silicon-b')"
                        },
                        "test_set": {
                            "type": "string",
                            "description": "Set di test da usare. Esempi: 'paraphrase' per test di parafrasi (PARA_*), 'standard' per test normali (TEST_*). Corrisponde al file tests_{name}.json. Default: 'standard'"
                        },
                        "mode": {
                            "type": "string",
                            "enum": ["auto", "assisted", "train"],
                            "description": "Modalita' di test (default: 'auto')",
                            "default": "auto"
                        },
                        "tests": {
                            "type": "string",
                            "enum": ["all", "pending", "failed"],
                            "description": "Quali test eseguire (default: 'pending')",
                            "default": "pending"
                        },
                        "new_run": {
                            "type": "boolean",
                            "description": "Se creare una nuova RUN su Google Sheets (default: false)",
                            "default": False
                        },
                        "test_limit": {
                            "type": "integer",
                            "description": "Limita a N test (0 = nessun limite)",
                            "default": 0
                        },
                        "test_ids": {
                            "type": "string",
                            "description": "Lista di ID test separati da virgola (es. 'TEST_001,TEST_002')"
                        },
                        "prompt_version": {
                            "type": "string",
                            "description": "Versione del prompt (es. 'v12'). Viene registrata su Google Sheets."
                        },
                        "parallel": {
                            "type": "boolean",
                            "description": "Abilita parallelismo per esecuzione veloce. Usa questo se l'utente chiede 'veloce', 'parallelo', 'il prima possibile'.",
                            "default": False
                        },
                        "repeat": {
                            "type": "integer",
                            "description": "Numero di RUN da creare (1-5). Usa questo se l'utente chiede 'lancia 2 run', '3 esecuzioni', etc. Ogni RUN crea un foglio separato su Google Sheets.",
                            "default": 1,
                            "minimum": 1,
                            "maximum": 5
                        },
                        "multi_testset": {
                            "type": "boolean",
                            "description": "Lancia 3 test set in parallelo (standard, paraphrase, GGP) con una singola pipeline. "
                                          "Usa questo quando l'utente chiede di lanciare 'tutti i test set', 'standard e paraphrase e ggp', "
                                          "'i 3 test set', etc. Ogni test set crea il proprio foglio Google Sheets.",
                            "default": False
                        },
                        "testsets": {
                            "type": "array",
                            "items": {"type": "string", "enum": ["standard", "paraphrase", "ggp"]},
                            "description": "Quali test set includere quando multi_testset=true. Default: tutti e 3.",
                            "default": ["standard", "paraphrase", "ggp"]
                        }
                    },
                    "required": ["project"]
                }
            ),
            Tool(
                name="get_pipeline_status",
                description="Ottiene lo stato delle pipeline CircleCI recenti",
                inputSchema={
                    "type": "object",
                    "properties": {
                        "limit": {
                            "type": "integer",
                            "description": "Numero massimo di pipeline da mostrare (default: 5)",
                            "default": 5
                        }
                    },
                    "required": []
                }
            ),
            Tool(
                name="get_workflow_status",
                description="Ottiene lo stato dettagliato di un workflow CircleCI",
                inputSchema={
                    "type": "object",
                    "properties": {
                        "pipeline_id": {
                            "type": "string",
                            "description": "ID della pipeline CircleCI"
                        }
                    },
                    "required": ["pipeline_id"]
                }
            ),
            Tool(
                name="list_runs",
                description="""OBBLIGATORIO per vedere le RUN disponibili su Google Sheets.

DEVI SEMPRE usare questo tool quando l'utente chiede:
- "Quali run ci sono?"
- "Mostrami le run di silicon-b"
- "Quante run abbiamo fatto?"
- "Lista run"

NON tentare di rispondere senza chiamare questo tool.
I dati delle RUN sono su Google Sheets e cambiano continuamente.""",
                inputSchema={
                    "type": "object",
                    "properties": {
                        "project": {
                            "type": "string",
                            "description": "Nome del progetto (es. 'silicon-b')"
                        }
                    },
                    "required": ["project"]
                }
            ),
            Tool(
                name="get_run_results",
                description="""OBBLIGATORIO per ottenere i risultati dei test da Google Sheets.

DEVI SEMPRE usare questo tool quando l'utente menziona:
- "risultati", "esiti", "come è andata"
- "mostrami la run", "vedi run"
- numeri di run (es. "run 38", "ultima run")
- "cosa c'è nel foglio"

NON tentare di rispondere senza chiamare questo tool.
I dati sono su Google Sheets e cambiano continuamente.

Parametro verbose=true per vedere anche le risposte del chatbot (CONVERSATION).""",
                inputSchema={
                    "type": "object",
                    "properties": {
                        "project": {
                            "type": "string",
                            "description": "Nome del progetto (es. 'silicon-b')"
                        },
                        "run_number": {
                            "type": "integer",
                            "description": "Numero della RUN (es. 15). Se non specificato, usa l'ultima."
                        },
                        "verbose": {
                            "type": "boolean",
                            "description": "Se true, mostra anche la risposta del chatbot (CONVERSATION). Default: false",
                            "default": False
                        }
                    },
                    "required": ["project"]
                }
            ),
            Tool(
                name="get_failed_tests",
                description="""OBBLIGATORIO per vedere i test falliti di una RUN.

DEVI SEMPRE usare questo tool quando l'utente chiede:
- "quali test sono falliti?"
- "mostrami i fallimenti"
- "errori della run"
- "test che non passano"

NON tentare di rispondere senza chiamare questo tool.
I dati sono su Google Sheets e cambiano continuamente.""",
                inputSchema={
                    "type": "object",
                    "properties": {
                        "project": {
                            "type": "string",
                            "description": "Nome del progetto (es. 'silicon-b')"
                        },
                        "run_number": {
                            "type": "integer",
                            "description": "Numero della RUN. Se non specificato, usa l'ultima."
                        }
                    },
                    "required": ["project"]
                }
            ),
            Tool(
                name="compare_runs",
                description="""OBBLIGATORIO per confrontare due RUN e trovare differenze.

DEVI SEMPRE usare questo tool quando l'utente chiede:
- "confronta la run 37 con la 38"
- "cosa è cambiato tra le run?"
- "mostrami le differenze"
- "regressioni tra run"
- "confronto run"

Mostra:
- Test che sono migliorati (FAIL → PASS)
- Test che sono peggiorati (PASS → FAIL)
- Statistiche comparative

NON tentare di rispondere senza chiamare questo tool.
I dati sono su Google Sheets e cambiano continuamente.""",
                inputSchema={
                    "type": "object",
                    "properties": {
                        "project": {
                            "type": "string",
                            "description": "Nome del progetto (es. 'silicon-b')"
                        },
                        "run_a": {
                            "type": "integer",
                            "description": "Numero della prima RUN (baseline)"
                        },
                        "run_b": {
                            "type": "integer",
                            "description": "Numero della seconda RUN (da confrontare). Se non specificato, usa l'ultima."
                        }
                    },
                    "required": ["project"]
                }
            ),
            # ====== NUOVI TOOL FOOL-PROOF ======
            Tool(
                name="start_test_session",
                description="""Avvia una sessione guidata per testare un progetto.

USA QUESTO TOOL quando l'utente dice:
- "Voglio testare silicon-b"
- "Lancia i test su silicon-b"
- "Esegui i test di parafrasi su silicon-b"
- "Testa il chatbot silicon-b"

IMPORTANTE: L'utente DEVE specificare il nome del progetto.
Se non lo fa, chiedi "Quale progetto vuoi testare?" e usa list_projects.

Il tool mostra:
1. Test set disponibili (es. standard, paraphrase, ggp) con conteggio
2. Stato attuale (test pending, falliti, pass rate ultima RUN)
3. Raccomandazione su cosa testare

DOPO aver mostrato le opzioni, CHIEDI all'utente quale test set vuole usare.""",
                inputSchema={
                    "type": "object",
                    "properties": {
                        "project": {
                            "type": "string",
                            "description": "Nome del progetto da testare (es. silicon-b, silicon-a)"
                        }
                    },
                    "required": ["project"]
                }
            ),
            Tool(
                name="prepare_test_run",
                description="""Prepara l'esecuzione dei test e mostra un RIEPILOGO DETTAGLIATO per conferma.

USA QUESTO TOOL quando l'utente ha scelto cosa testare:
- "Ok, lancia i test standard"
- "Vai con paraphrase"
- "Esegui i pending"

NON lancia ancora! Mostra il riepilogo e CHIEDI CONFERMA all'utente.
Solo dopo che l'utente conferma (es. "sì", "procedi", "ok"), usa execute_test_run.""",
                inputSchema={
                    "type": "object",
                    "properties": {
                        "project": {
                            "type": "string",
                            "description": "Nome del progetto"
                        },
                        "test_set": {
                            "type": "string",
                            "description": "Test set (standard, paraphrase, ggp)"
                        },
                        "tests": {
                            "type": "string",
                            "enum": ["pending", "all", "failed"],
                            "description": "Quali test eseguire",
                            "default": "pending"
                        }
                    },
                    "required": ["project", "test_set"]
                }
            ),
            Tool(
                name="execute_test_run",
                description="""Esegue i test DOPO che l'utente ha confermato il riepilogo.

USA QUESTO TOOL SOLO quando l'utente ha visto il riepilogo di prepare_test_run
e ha confermato esplicitamente:
- "sì"
- "procedi"
- "ok vai"
- "confermo"

Se l'utente NON ha confermato o ha detto "no", NON usare questo tool.""",
                inputSchema={
                    "type": "object",
                    "properties": {
                        "project": {
                            "type": "string",
                            "description": "Nome del progetto"
                        },
                        "test_set": {
                            "type": "string",
                            "description": "Test set (standard, paraphrase, ggp)"
                        },
                        "tests": {
                            "type": "string",
                            "enum": ["pending", "all", "failed"],
                            "description": "Quali test eseguire",
                            "default": "pending"
                        }
                    },
                    "required": ["project", "test_set"]
                }
            ),
            Tool(
                name="check_pipeline_status",
                description="""Controlla lo stato di una pipeline in esecuzione.

USA QUESTO TOOL quando l'utente chiede:
- "Come sta andando?"
- "A che punto siamo?"
- "Sono finiti i test?"
- "Controlla lo stato"

NON fa polling automatico (non possibile in MCP).
Restituisce lo stato attuale + promemoria per ricontrollare.""",
                inputSchema={
                    "type": "object",
                    "properties": {
                        "pipeline_id": {
                            "type": "string",
                            "description": "ID pipeline da controllare"
                        }
                    },
                    "required": ["pipeline_id"]
                }
            ),
            Tool(
                name="show_results",
                description="""Mostra i risultati dell'ultima esecuzione in modo chiaro e comprensibile.

USA QUESTO TOOL quando l'utente chiede:
- "Come è andata?"
- "Mostrami i risultati di silicon-b"
- "Quanti test sono passati?"
- "Ci sono errori?"

Mostra:
- Pass rate con indicatore visivo (✅ >90%, ⚠️ 70-90%, ❌ <70%)
- Numero test passati/falliti/pending
- Lista test falliti con breve descrizione
- Suggerimento per prossimi passi""",
                inputSchema={
                    "type": "object",
                    "properties": {
                        "project": {
                            "type": "string",
                            "description": "Nome del progetto"
                        },
                        "run_number": {
                            "type": "integer",
                            "description": "Numero RUN (se non specificato, usa l'ultima)"
                        }
                    },
                    "required": ["project"]
                }
            ),
            # Tool per aggiungere nuovi test
            Tool(
                name="add_test",
                description="""Aggiunge un nuovo test al test set di un progetto.

USA QUESTO TOOL quando l'utente chiede:
- "Aggiungi questo test a silicon-b"
- "Crea un nuovo test per..."
- "Voglio aggiungere una domanda al test set"

Il test verrà aggiunto al file tests_{test_set}.json del progetto.
Genera automaticamente un ID univoco basato sul test set (es. TEST_055, PARA_025).""",
                inputSchema={
                    "type": "object",
                    "properties": {
                        "project": {
                            "type": "string",
                            "description": "Nome del progetto (es. 'silicon-b')"
                        },
                        "test_set": {
                            "type": "string",
                            "description": "Set di test dove aggiungere (standard, paraphrase, ggp). Default: 'standard'",
                            "default": "standard"
                        },
                        "question": {
                            "type": "string",
                            "description": "La domanda da testare"
                        },
                        "expected_answer": {
                            "type": "string",
                            "description": "Risposta attesa (opzionale, per valutazione semantica)"
                        },
                        "category": {
                            "type": "string",
                            "description": "Categoria del test (opzionale)"
                        },
                        "follow_ups": {
                            "type": "array",
                            "items": {"type": "string"},
                            "description": "Lista di domande follow-up (opzionale)"
                        }
                    },
                    "required": ["project", "question"]
                }
            ),
            # Tool per notificare Corrado via Telegram
            Tool(
                name="notify_corrado",
                description="""Invia una notifica a Corrado via Telegram.

USA QUESTO TOOL quando l'utente dice:
- "Avvisa Corrado che..."
- "Notifica a Corrado..."
- "Manda un messaggio a Corrado..."
- "Segnala a Corrado..."
- "Ho un problema, avvisa Corrado"

Il messaggio arriva istantaneamente su Telegram con contesto sul progetto.""",
                inputSchema={
                    "type": "object",
                    "properties": {
                        "message": {
                            "type": "string",
                            "description": "Il messaggio da inviare a Corrado"
                        },
                        "project": {
                            "type": "string",
                            "description": "Progetto correlato (opzionale, per contesto)"
                        },
                        "priority": {
                            "type": "string",
                            "enum": ["low", "normal", "high"],
                            "description": "Priorità della notifica",
                            "default": "normal"
                        }
                    },
                    "required": ["message"]
                }
            ),
            # Tool per vedere le novità
            Tool(
                name="novita",
                description="""Mostra le ultime novità e funzionalità aggiunte.

USA QUESTO TOOL quando l'utente chiede:
- "Cosa c'è di nuovo?"
- "Novità"
- "Ultime modifiche"
- "Changelog"
- "Cosa è cambiato?"

Mostra le ultime versioni con le modifiche principali.""",
                inputSchema={
                    "type": "object",
                    "properties": {
                        "limit": {
                            "type": "integer",
                            "description": "Numero di versioni da mostrare (default: 3)",
                            "default": 3
                        }
                    },
                    "required": []
                }
            ),
            # ====== NUOVI TOOL v1.4.0 ======
            Tool(
                name="detect_flaky_tests",
                description="""Rileva test flaky (risultati inconsistenti) nelle ultime N run.

USA QUESTO TOOL quando l'utente chiede:
- "Quali test sono instabili?"
- "Test flaky"
- "Test che a volte passano e a volte falliscono"
- "Stabilità dei test"

Un test è flaky se ha risultati inconsistenti (PASS/FAIL) su run diverse.""",
                inputSchema={
                    "type": "object",
                    "properties": {
                        "project": {
                            "type": "string",
                            "description": "Nome del progetto (es. 'silicon-b')"
                        },
                        "last_n_runs": {
                            "type": "integer",
                            "description": "Numero RUN da analizzare (default: 10)",
                            "default": 10
                        },
                        "threshold": {
                            "type": "number",
                            "description": "Soglia flaky score 0-1 (0=stabile, 1=random, default: 0.3)",
                            "default": 0.3
                        }
                    },
                    "required": ["project"]
                }
            ),
            Tool(
                name="get_regressions",
                description="""Rileva test che erano PASS e ora sono FAIL (regressioni).

USA QUESTO TOOL quando l'utente chiede:
- "Quali test sono peggiorati?"
- "Regressioni nell'ultima run"
- "Test che prima passavano"
- "Cosa si è rotto?"

Confronta l'ultima RUN con quella precedente automaticamente.""",
                inputSchema={
                    "type": "object",
                    "properties": {
                        "project": {
                            "type": "string",
                            "description": "Nome del progetto"
                        },
                        "run_number": {
                            "type": "integer",
                            "description": "RUN da verificare (default: ultima)"
                        }
                    },
                    "required": ["project"]
                }
            ),
            Tool(
                name="get_performance_report",
                description="""Genera report dettagliato delle performance di una RUN.

USA QUESTO TOOL quando l'utente chiede:
- "Come sono andate le performance?"
- "Quanto ci ha messo?"
- "Velocità dei test"
- "Metriche performance"

Mostra timing, throughput, latenze servizi esterni.""",
                inputSchema={
                    "type": "object",
                    "properties": {
                        "project": {
                            "type": "string",
                            "description": "Nome del progetto"
                        },
                        "run_number": {
                            "type": "integer",
                            "description": "RUN da analizzare (default: ultima)"
                        }
                    },
                    "required": ["project"]
                }
            ),
            Tool(
                name="get_performance_alerts",
                description="""Verifica se ci sono alert di performance (soglie superate).

USA QUESTO TOOL quando l'utente chiede:
- "Ci sono problemi di performance?"
- "Alert?"
- "Soglie superate?"
- "Warning performance"

Controlla error rate, pass rate, latenze e confronta con baseline.""",
                inputSchema={
                    "type": "object",
                    "properties": {
                        "project": {
                            "type": "string",
                            "description": "Nome del progetto"
                        },
                        "run_number": {
                            "type": "integer",
                            "description": "RUN da verificare (default: ultima)"
                        }
                    },
                    "required": ["project"]
                }
            ),
            Tool(
                name="analyze_coverage",
                description="""Analizza la copertura dei test per categoria.

USA QUESTO TOOL quando l'utente chiede:
- "Quali categorie sono coperte?"
- "Gap di copertura"
- "Mancano test?"
- "Coverage dei test"

Identifica categorie scoperte o con pochi test.""",
                inputSchema={
                    "type": "object",
                    "properties": {
                        "project": {
                            "type": "string",
                            "description": "Nome del progetto"
                        },
                        "test_set": {
                            "type": "string",
                            "description": "Test set da analizzare (default: 'standard')",
                            "default": "standard"
                        }
                    },
                    "required": ["project"]
                }
            ),
            Tool(
                name="get_stability_report",
                description="""Report generale sulla stabilità della suite di test.

USA QUESTO TOOL quando l'utente chiede:
- "Quanto è stabile la suite?"
- "Stabilità complessiva"
- "Overview stabilità"
- "Report stabilità"

Mostra test stabili, instabili e flaky con score complessivo.""",
                inputSchema={
                    "type": "object",
                    "properties": {
                        "project": {
                            "type": "string",
                            "description": "Nome del progetto"
                        },
                        "last_n_runs": {
                            "type": "integer",
                            "description": "Numero RUN da analizzare (default: 10)",
                            "default": 10
                        }
                    },
                    "required": ["project"]
                }
            ),
            Tool(
                name="diagnose_prompt",
                description="""Diagnosi intelligente dei test falliti con suggerimenti fix.

USA QUESTO TOOL quando l'utente chiede:
- "Perché questo test fallisce?"
- "Come fixo questo errore?"
- "Diagnosi fallimenti"
- "Suggerimenti per il prompt"

Classifica il tipo di errore e suggerisce fix specifici.""",
                inputSchema={
                    "type": "object",
                    "properties": {
                        "project": {
                            "type": "string",
                            "description": "Nome del progetto"
                        },
                        "run_number": {
                            "type": "integer",
                            "description": "RUN da analizzare (default: ultima)"
                        },
                        "test_id": {
                            "type": "string",
                            "description": "ID specifico test da diagnosticare (opzionale)"
                        }
                    },
                    "required": ["project"]
                }
            ),
            Tool(
                name="calibrate_thresholds",
                description="""Analizza metriche storiche e suggerisce soglie ottimali.

USA QUESTO TOOL quando l'utente chiede:
- "Quali soglie usare?"
- "Calibra metriche"
- "Soglie ottimali"
- "Configura thresholds"

Analizza distribuzioni e suggerisce configurazione ottimale.""",
                inputSchema={
                    "type": "object",
                    "properties": {
                        "project": {
                            "type": "string",
                            "description": "Nome del progetto"
                        },
                        "last_n_runs": {
                            "type": "integer",
                            "description": "Numero RUN da analizzare (default: 10)",
                            "default": 10
                        }
                    },
                    "required": ["project"]
                }
            ),
            Tool(
                name="export_report",
                description="""Esporta i risultati di una RUN in vari formati.

USA QUESTO TOOL quando l'utente chiede:
- "Esporta in Excel"
- "Genera report HTML"
- "Scarica CSV"
- "Export risultati"

Formati disponibili: excel, html, csv.""",
                inputSchema={
                    "type": "object",
                    "properties": {
                        "project": {
                            "type": "string",
                            "description": "Nome del progetto"
                        },
                        "run_number": {
                            "type": "integer",
                            "description": "RUN da esportare (default: ultima)"
                        },
                        "format": {
                            "type": "string",
                            "enum": ["excel", "html", "csv"],
                            "description": "Formato export (excel, html, csv)",
                            "default": "excel"
                        }
                    },
                    "required": ["project"]
                }
            ),
            Tool(
                name="debug_trace",
                description="""Analizza un trace LangSmith per debug di test falliti o con esito negativo.

USA QUESTO TOOL quando l'utente chiede:
- "Analizza il trace xyz"
- "Perché questo trace ha fallito?"
- "Debug trace abc-123"
- "Cosa è successo nel trace?"
- "Fammi vedere il trace"

Fetcha il trace completo da LangSmith e mostra:
- Input/output della conversazione
- Tool chiamati e loro output
- Timing breakdown
- Errori e problemi

Il trace_id si trova nella colonna LS TRACE URL dei risultati su Google Sheets.""",
                inputSchema={
                    "type": "object",
                    "properties": {
                        "trace_id": {
                            "type": "string",
                            "description": "ID del trace LangSmith (es. abc123-def456-...)"
                        },
                        "project": {
                            "type": "string",
                            "description": "Progetto per le credenziali LangSmith (default: silicon-b)",
                            "default": "silicon-b"
                        }
                    },
                    "required": ["trace_id"]
                }
            ),
        ]
        logger.info(f"=== Returning {len(tools)} tools ===")
        return tools

    @server.call_tool()
    async def call_tool(name: str, arguments: dict) -> list[TextContent]:
        """Handle tool calls."""
        logger.info(f"Tool call: {name} with args: {arguments}")

        try:
            if name == "get_help":
                return await handle_get_help(arguments)
            elif name == "list_projects":
                return await handle_list_projects()
            elif name == "suggest_project":
                return await handle_suggest_project()
            elif name == "list_test_sets":
                return await handle_list_test_sets(arguments)
            elif name == "list_tests":
                return await handle_list_tests(arguments)
            elif name == "trigger_circleci":
                return await handle_trigger_circleci(arguments)
            elif name == "get_pipeline_status":
                return await handle_get_pipeline_status(arguments)
            elif name == "get_workflow_status":
                return await handle_get_workflow_status(arguments)
            elif name == "list_runs":
                return await handle_list_runs(arguments)
            elif name == "get_run_results":
                return await handle_get_run_results(arguments)
            elif name == "get_failed_tests":
                return await handle_get_failed_tests(arguments)
            elif name == "compare_runs":
                return await handle_compare_runs(arguments)
            # ====== NUOVI TOOL FOOL-PROOF ======
            elif name == "start_test_session":
                return await handle_start_test_session(arguments)
            elif name == "prepare_test_run":
                return await handle_prepare_test_run(arguments)
            elif name == "execute_test_run":
                return await handle_execute_test_run(arguments)
            elif name == "check_pipeline_status":
                return await handle_check_pipeline_status(arguments)
            elif name == "show_results":
                return await handle_show_results(arguments)
            elif name == "add_test":
                return await handle_add_test(arguments)
            elif name == "notify_corrado":
                return await handle_notify_corrado(arguments)
            elif name == "novita":
                return await handle_novita(arguments)
            # ====== NUOVI TOOL v1.4.0 ======
            elif name == "detect_flaky_tests":
                return await handle_detect_flaky_tests(arguments)
            elif name == "get_regressions":
                return await handle_get_regressions(arguments)
            elif name == "get_performance_report":
                return await handle_get_performance_report(arguments)
            elif name == "get_performance_alerts":
                return await handle_get_performance_alerts(arguments)
            elif name == "analyze_coverage":
                return await handle_analyze_coverage(arguments)
            elif name == "get_stability_report":
                return await handle_get_stability_report(arguments)
            elif name == "diagnose_prompt":
                return await handle_diagnose_prompt(arguments)
            elif name == "calibrate_thresholds":
                return await handle_calibrate_thresholds(arguments)
            elif name == "export_report":
                return await handle_export_report(arguments)
            elif name == "debug_trace":
                return await handle_debug_trace(arguments)
            else:
                return [TextContent(
                    type="text",
                    text=f"""Tool sconosciuto: {name}

💡 *Questa funzionalità non esiste ancora.*
Vuoi che la richieda? Scrivi: **"Avvisa Corrado che vorrei poter {name}"**"""
                )]
        except Exception as e:
            logger.error(f"Error in tool {name}: {e}")
            return [TextContent(
                type="text",
                text=f"""Errore nell'esecuzione del tool {name}: {str(e)}

💡 *Hai bisogno di aiuto?*
Scrivi: **"Avvisa Corrado che ho un problema con {name}"**"""
            )]


async def handle_get_help(arguments: dict) -> list[TextContent]:
    """Handle get_help tool."""
    topic = arguments.get("topic", "all")

    help_sections = {
        "overview": """## Chatbot Tester - Panoramica

Chatbot Tester è un sistema per testare automaticamente chatbot conversazionali.

**Progetti disponibili:** Usa `list_projects` per vedere i progetti configurati.

**Flusso tipico:**
1. Scegli un progetto (es. `silicon-b`)
2. Esegui i test con `trigger_circleci`
3. Controlla i risultati con `list_runs` e `get_run_results`
4. Analizza regressioni con `compare_runs`
""",

        "testing": """## Come eseguire i test

### Comando base
> "Esegui i test pending di silicon-b"

### Parametri disponibili per `trigger_circleci`:

| Parametro | Valori | Default | Descrizione |
|-----------|--------|---------|-------------|
| `project` | nome progetto | - | **Obbligatorio** |
| `mode` | auto, assisted, train | auto | Modalità esecuzione |
| `tests` | all, pending, failed | pending | Quali test eseguire |
| `new_run` | true/false | false | Crea nuova RUN su Sheets |
| `test_limit` | numero | 0 (tutti) | Limita numero test |
| `test_ids` | "ID1,ID2,..." | - | Test specifici |

### Esempi:
- *"Esegui tutti i test di silicon-b con nuova RUN"*
- *"Esegui solo i test falliti di silicon-b"*
- *"Esegui TEST_001 e TEST_050 su silicon-b"*
- *"Esegui 5 test di silicon-b"*

### Modalità:
- **auto**: Esecuzione automatica, valutazione AI
- **assisted**: Richiede conferma manuale
- **train**: Per generare dati di training
""",

        "results": """## Come vedere i risultati

### Lista RUN disponibili
> "Mostra le RUN di silicon-b"

### Risultati di una RUN specifica
> "Mostra i risultati della RUN 15 di silicon-b"

### Solo test falliti
> "Mostra i test falliti della RUN 15"

### Stato pipeline CircleCI
> "Mostra lo stato delle pipeline recenti"
> "Dettaglio della pipeline xyz123"

### Informazioni mostrate:
- Pass rate (% test passati)
- Lista test con esito (PASS/FAIL)
- Timestamp esecuzione
- Ambiente (DEV/STAGING/PROD)
""",

        "comparison": """## Come confrontare RUN

### Confronto tra due RUN
> "Confronta le RUN 10 e 15 di silicon-b"

### Cosa mostra il confronto:
- **Regressioni**: Test che erano PASS e ora sono FAIL
- **Miglioramenti**: Test che erano FAIL e ora sono PASS
- **Delta pass rate**: Variazione percentuale

### Quando usarlo:
- Dopo un deploy per verificare regressioni
- Per confrontare performance tra ambienti
- Per tracking qualità nel tempo
""",

        "quickstart": """## 🚀 Guida Rapida

**Se non sai da dove partire, dimmi semplicemente:**
- *"Voglio testare silicon-b"* → Ti guido io passo passo
- *"Mostrami cosa posso fare"* → Lista completa opzioni

### Test Set disponibili per silicon-b:
- **standard**: 50+ test completi
- **paraphrase**: 40+ test con domande riformulate
- **ggp**: 4 test di grounding

### Comandi rapidi:
- *"Lancia i test pending di silicon-b"* → Esegue solo test non ancora fatti
- *"Lancia tutti i test"* → Nuova RUN completa
- *"Come sta andando?"* → Stato pipeline
- *"Mostrami i risultati"* → Ultima RUN

### Flusso tipico:
1. **Scegli progetto**: "Voglio testare silicon-b"
2. **Scegli test set**: "standard" / "paraphrase" / "ggp"
3. **Conferma**: Vedi riepilogo e confermi
4. **Monitora**: "Come sta andando?"
5. **Risultati**: "Mostrami i risultati"

---
💡 *Scrivi "novità" per vedere le ultime funzionalità aggiunte!*
🔔 *Problemi? Scrivi "Avvisa Corrado che..." per notificare il team*
"""
    }

    if topic == "all":
        result = "# Guida Chatbot Tester\n\n"
        result += help_sections["overview"] + "\n---\n\n"
        result += help_sections["testing"] + "\n---\n\n"
        result += help_sections["results"] + "\n---\n\n"
        result += help_sections["comparison"]
    else:
        result = help_sections.get(topic, "Argomento non trovato. Usa: overview, testing, results, comparison, all")

    return [TextContent(type="text", text=result)]


async def handle_list_projects() -> list[TextContent]:
    """Handle list_projects tool."""
    projects = get_available_projects()

    if not projects:
        return [TextContent(
            type="text",
            text="Nessun progetto trovato."
        )]

    result = "Progetti disponibili:\n\n"
    for project in projects:
        test_sets = get_available_test_sets(project)
        total = sum(test_sets.values())
        sets_info = ", ".join([f"{name}: {count}" for name, count in test_sets.items()])
        result += f"- **{project}**: {total} test ({sets_info})\n"

    return [TextContent(type="text", text=result)]


async def handle_list_test_sets(arguments: dict) -> list[TextContent]:
    """Handle list_test_sets tool."""
    project = arguments.get("project")

    if not project:
        return [TextContent(type="text", text="Errore: specificare il nome del progetto")]

    if project not in get_available_projects():
        return [TextContent(
            type="text",
            text=f"Progetto '{project}' non trovato."
        )]

    test_sets = get_available_test_sets(project)

    if not test_sets:
        return [TextContent(type="text", text=f"Nessun set di test trovato per '{project}'.")]

    result = f"## Set di test disponibili per {project}\n\n"

    for name, count in sorted(test_sets.items()):
        result += f"- **{name}**: {count} test\n"

    result += "\n---\n"
    result += "Per usare un set specifico:\n"
    result += f"*\"Esegui i test paraphrase di {project}\"*\n"
    result += f"*\"Mostra i test paraphrase di {project}\"*"

    return [TextContent(type="text", text=result)]


async def handle_suggest_project() -> list[TextContent]:
    """Handle suggest_project tool - analyzes projects and suggests which to test."""
    projects = get_available_projects()

    if not projects:
        return [TextContent(type="text", text="Nessun progetto trovato.")]

    project_stats = []

    for project in projects:
        tests = get_project_tests(project)
        total_tests = len(tests)

        stats = {
            "name": project,
            "total_tests": total_tests,
            "last_run": None,
            "last_run_date": None,
            "pass_rate": None,
            "pending": 0,
            "failed": 0,
            "priority_score": 0,
            "priority_reason": []
        }

        # Try to get last RUN info
        try:
            client = get_sheets_client(project)
            runs = client.get_all_run_numbers()

            if runs:
                last_run = max(runs)
                stats["last_run"] = last_run

                # Get last run results
                client.active_run = last_run
                results = client.get_all_results()

                if results:
                    passed = sum(1 for r in results if get_field(r, "esito").upper() == "PASS")
                    failed = sum(1 for r in results if get_field(r, "esito").upper() == "FAIL")
                    pending = total_tests - passed - failed

                    stats["pass_rate"] = (passed / len(results) * 100) if results else 0
                    stats["pending"] = pending
                    stats["failed"] = failed

                    # Get run info for date
                    run_info = client.get_run_info(last_run)
                    if run_info:
                        stats["last_run_date"] = run_info.get("timestamp", "?")

        except Exception as e:
            logger.warning(f"Could not get stats for {project}: {e}")

        # Calculate priority score
        score = 0

        # High priority: has pending tests
        if stats["pending"] > 0:
            score += 30
            stats["priority_reason"].append(f"{stats['pending']} test pending")

        # High priority: has failed tests
        if stats["failed"] > 0:
            score += 20
            stats["priority_reason"].append(f"{stats['failed']} test falliti")

        # Medium priority: low pass rate
        if stats["pass_rate"] is not None and stats["pass_rate"] < 80:
            score += 15
            stats["priority_reason"].append(f"pass rate {stats['pass_rate']:.0f}%")

        # Low priority: never run
        if stats["last_run"] is None:
            score += 10
            stats["priority_reason"].append("mai eseguito")

        # Base score from test count (more tests = slightly higher priority)
        score += min(stats["total_tests"] / 10, 5)

        stats["priority_score"] = score
        project_stats.append(stats)

    # Sort by priority score (descending)
    project_stats.sort(key=lambda x: x["priority_score"], reverse=True)

    # Build result
    result = "## Analisi Progetti\n\n"

    for i, stats in enumerate(project_stats):
        is_suggested = (i == 0)
        prefix = "**>>> " if is_suggested else ""
        suffix = " <<<**" if is_suggested else ""

        result += f"{prefix}{stats['name']}{suffix}\n"
        result += f"- Test totali: {stats['total_tests']}\n"

        if stats["last_run"]:
            result += f"- Ultima RUN: #{stats['last_run']}"
            if stats["last_run_date"]:
                result += f" ({stats['last_run_date']})"
            result += "\n"

            if stats["pass_rate"] is not None:
                result += f"- Pass rate: {stats['pass_rate']:.0f}%\n"
            if stats["pending"] > 0:
                result += f"- Pending: {stats['pending']}\n"
            if stats["failed"] > 0:
                result += f"- Falliti: {stats['failed']}\n"
        else:
            result += "- *Nessuna RUN eseguita*\n"

        result += "\n"

    # Add suggestion
    suggested = project_stats[0]
    result += "---\n\n"
    result += f"## Suggerimento\n\n"
    result += f"**Consiglio di testare `{suggested['name']}`**"

    if suggested["priority_reason"]:
        result += f" perché: {', '.join(suggested['priority_reason'])}"

    result += f"\n\nPer eseguire: *\"Esegui i test pending di {suggested['name']}\"*"

    return [TextContent(type="text", text=result)]


async def handle_list_tests(arguments: dict) -> list[TextContent]:
    """Handle list_tests tool."""
    project = arguments.get("project")
    test_set = arguments.get("test_set", "standard")
    filter_text = arguments.get("filter", "").lower()

    if not project:
        return [TextContent(
            type="text",
            text="Errore: specificare il nome del progetto"
        )]

    if project not in get_available_projects():
        return [TextContent(
            type="text",
            text=f"Progetto '{project}' non trovato. Usa list_projects per vedere i progetti disponibili."
        )]

    tests = get_project_tests(project, test_set)

    if not tests:
        return [TextContent(
            type="text",
            text=f"Nessun test trovato per il progetto '{project}'."
        )]

    # Apply filter if provided
    if filter_text:
        filtered_tests = []
        for test in tests:
            test_id = test.get("id", "").lower()
            description = test.get("description", "").lower()
            question = test.get("question", "").lower()
            category = test.get("category", "").lower()

            if (filter_text in test_id or
                filter_text in description or
                filter_text in question or
                filter_text in category):
                filtered_tests.append(test)

        if not filtered_tests:
            return [TextContent(
                type="text",
                text=f"Nessun test trovato con filtro '{filter_text}' nel progetto '{project}'."
            )]

        result = f"Test per **{project}** filtrati per '{filter_text}' ({len(filtered_tests)} trovati):\n\n"

        # List all filtered tests (no category grouping for filtered results)
        for test in filtered_tests:
            test_id = test.get("id", "?")
            description = test.get("description", test.get("question", ""))[:80]
            result += f"- `{test_id}`: {description}\n"

        # Add hint for running these tests
        if len(filtered_tests) <= 10:
            test_ids = ",".join([t.get("id", "") for t in filtered_tests if t.get("id")])
            result += f"\n---\nPer eseguire questi test:\n"
            result += f"*\"Esegui i test {test_ids} su {project}\"*"

        return [TextContent(type="text", text=result)]

    # No filter - show ALL tests grouped by category
    result = f"Test per **{project}** - set **{test_set}** ({len(tests)} totali):\n\n"

    # Group by category if available
    by_category = {}
    for test in tests:
        category = test.get("category", "Uncategorized")
        if category not in by_category:
            by_category[category] = []
        by_category[category].append(test)

    for category, cat_tests in sorted(by_category.items()):
        result += f"### {category} ({len(cat_tests)} test)\n"
        for test in cat_tests:  # Show ALL tests, no limit
            test_id = test.get("id", "?")
            question = test.get("question", test.get("description", ""))[:80]
            result += f"- `{test_id}`: {question}\n"
        result += "\n"

    return [TextContent(type="text", text=result)]


async def handle_trigger_circleci(arguments: dict) -> list[TextContent]:
    """Handle trigger_circleci tool."""
    from src.circleci_client import CircleCIClient

    project = arguments.get("project")
    test_set = arguments.get("test_set", "standard")
    mode = arguments.get("mode", "auto")
    tests = arguments.get("tests", "pending")
    new_run = arguments.get("new_run", False)
    test_limit = arguments.get("test_limit", 0)
    test_ids = arguments.get("test_ids", "")
    prompt_version = arguments.get("prompt_version", "")
    parallel = arguments.get("parallel", False)
    repeat = arguments.get("repeat", 1)
    multi_testset = arguments.get("multi_testset", False)
    testsets = arguments.get("testsets", ["standard", "paraphrase", "ggp"])

    # Se parallel=True, usa 3 container CircleCI per velocizzare
    native_parallelism = 3 if parallel else 0

    # Valida repeat
    repeat = max(1, min(5, repeat))

    if not project:
        return [TextContent(
            type="text",
            text="Errore: specificare il nome del progetto"
        )]

    # Validate project exists
    if project not in get_available_projects():
        return [TextContent(
            type="text",
            text=f"Progetto '{project}' non trovato. Usa list_projects per vedere i progetti disponibili."
        )]

    # Initialize CircleCI client early to cancel previous pipelines
    client = CircleCIClient()
    if not client.is_available():
        return [TextContent(
            type="text",
            text="Errore: CircleCI non configurato. Assicurarsi che CIRCLECI_TOKEN sia impostato."
        )]

    # Cancel any active/pending pipelines before starting a new one
    cancelled_count = 0
    try:
        pipelines = client.list_pipelines(limit=10)
        for pipeline in pipelines:
            if pipeline.is_active:
                workflows = client.get_pipeline_workflows(pipeline.id)
                for workflow in workflows:
                    if workflow.is_active:
                        if client.cancel_workflow(workflow.id):
                            cancelled_count += 1
                            logger.info(f"Cancelled workflow {workflow.id} from pipeline #{pipeline.number}")
    except Exception as e:
        logger.warning(f"Error cancelling previous pipelines: {e}")

    # Se multi_testset=True, usa il workflow dedicato
    if multi_testset:
        testset_standard = "standard" in testsets
        testset_paraphrase = "paraphrase" in testsets
        testset_ggp = "ggp" in testsets

        success, response = client.trigger_pipeline(
            project=project,
            mode=mode,
            tests=tests,
            new_run=True,  # Sempre new_run per multi_testset
            test_limit=test_limit,
            prompt_version=prompt_version,
            multi_testset=True,
            testset_standard=testset_standard,
            testset_paraphrase=testset_paraphrase,
            testset_ggp=testset_ggp
        )

        if not success:
            error = response.get("error", "Errore sconosciuto") if response else "Errore sconosciuto"
            return [TextContent(type="text", text=f"Errore nell'avvio della pipeline: {error}")]

        pipeline_id = response.get("id", "?")
        pipeline_number = response.get("number", "?")
        pipeline_url = f"https://app.circleci.com/pipelines/gh/corradofrancolini/chatbot-tester-private/{pipeline_number}"

        active_testsets = []
        if testset_standard:
            active_testsets.append("standard (→ Run NNN)")
        if testset_paraphrase:
            active_testsets.append("paraphrase (→ PAR NNN)")
        if testset_ggp:
            active_testsets.append("ggp (→ GGP NNN)")

        cancelled_msg = f"\n🛑 **{cancelled_count} workflow precedenti cancellati.**\n" if cancelled_count > 0 else ""

        result = f"""✅ **Pipeline Multi-Testset avviata!**
{cancelled_msg}
**Dettagli:**
- Pipeline ID: `{pipeline_id}`
- Numero: #{pipeline_number}
- Progetto: {project}
- Modalita': {mode}

**Test set in esecuzione (3 job paralleli):**
{chr(10).join(f'  • {ts}' for ts in active_testsets)}

**Ogni test set crea il proprio foglio Google Sheets.**

**Link:** {pipeline_url}

Usa `get_workflow_status` con pipeline_id `{pipeline_id}` per monitorare lo stato."""

        return [TextContent(type="text", text=result)]

    # If a non-standard test_set is specified, use the tests_file parameter instead of test_ids
    tests_file = "tests.json"
    if test_set != "standard" and not test_ids:
        tests_file = f"tests_{test_set}.json"
        tests = "all"  # Run all tests from the file
        logger.info(f"Using test set file '{tests_file}'")

    # Se repeat > 1, lancia multiple pipeline
    pipelines_launched = []
    for i in range(repeat):
        success, response = client.trigger_pipeline(
            project=project,
            mode=mode,
            tests=tests,
            new_run=new_run,
            test_limit=test_limit,
            test_ids=test_ids,
            tests_file=tests_file,
            prompt_version=prompt_version,
            native_parallelism=native_parallelism
        )
        if success:
            pipelines_launched.append({
                "id": response.get("id", "?"),
                "number": response.get("number", "?")
            })
        else:
            error = response.get("error", "Errore sconosciuto") if response else "Errore sconosciuto"
            if pipelines_launched:
                # Alcune pipeline sono partite, riporta comunque
                break
            else:
                return [TextContent(
                    type="text",
                    text=f"Errore nell'avvio della pipeline: {error}"
                )]

    cancelled_msg = f"\n🛑 **{cancelled_count} workflow precedenti cancellati.**\n" if cancelled_count > 0 else ""

    if len(pipelines_launched) == 1:
        # Singola pipeline
        p = pipelines_launched[0]
        pipeline_url = f"https://app.circleci.com/pipelines/gh/corradofrancolini/chatbot-tester-private/{p['number']}"

        result = f"""✅ **Pipeline avviata con successo!**
{cancelled_msg}
**Dettagli:**
- Pipeline ID: `{p['id']}`
- Numero: #{p['number']}
- Progetto: {project}
- Modalita': {mode}
- Test: {tests}
- Nuova RUN: {'Si' if new_run else 'No'}
{f'- Test set: {test_set}' if test_set != 'standard' else ''}
{f'- Prompt version: {prompt_version}' if prompt_version else ''}
{f'- Parallelismo: {native_parallelism} container' if native_parallelism > 0 else ''}
{f'- Limite test: {test_limit}' if test_limit > 0 else ''}
{f'- Test specifici: {test_ids}' if test_ids else ''}

**Link:** {pipeline_url}

Usa `get_workflow_status` con pipeline_id `{p['id']}` per monitorare lo stato."""
    else:
        # Multiple pipeline
        result = f"""✅ **{len(pipelines_launched)} Pipeline avviate con successo!**
{cancelled_msg}
**Configurazione comune:**
- Progetto: {project}
- Modalita': {mode}
- Test: {tests}
- Nuova RUN: Si (una per pipeline)
{f'- Test set: {test_set}' if test_set != 'standard' else ''}
{f'- Prompt version: {prompt_version}' if prompt_version else ''}
{f'- Parallelismo: {native_parallelism} container' if native_parallelism > 0 else ''}

**Pipeline lanciate:**
"""
        for i, p in enumerate(pipelines_launched, 1):
            url = f"https://app.circleci.com/pipelines/gh/corradofrancolini/chatbot-tester-private/{p['number']}"
            result += f"{i}. **#{p['number']}** - {url}\n"

        result += f"\nOgni pipeline creerà una RUN separata su Google Sheets per confronto A/B."

    return [TextContent(type="text", text=result)]


async def handle_get_pipeline_status(arguments: dict) -> list[TextContent]:
    """Handle get_pipeline_status tool."""
    from src.circleci_client import CircleCIClient

    limit = arguments.get("limit", 5)

    client = CircleCIClient()

    if not client.is_available():
        return [TextContent(
            type="text",
            text="Errore: CircleCI non configurato. Assicurarsi che CIRCLECI_TOKEN sia impostato."
        )]

    pipelines = client.list_pipelines(limit=limit)

    if not pipelines:
        return [TextContent(
            type="text",
            text="Nessuna pipeline trovata."
        )]

    result = f"Ultime {len(pipelines)} pipeline:\n\n"

    for p in pipelines:
        created = p.created_at[:19].replace("T", " ") if p.created_at else "?"
        result += f"{p.status_icon} **#{p.number}** [{p.state}] - {created}\n"
        result += f"   ID: `{p.id}`\n"
        result += f"   URL: {p.url}\n\n"

    return [TextContent(type="text", text=result)]


async def handle_get_workflow_status(arguments: dict) -> list[TextContent]:
    """Handle get_workflow_status tool."""
    from src.circleci_client import CircleCIClient

    pipeline_id = arguments.get("pipeline_id")

    if not pipeline_id:
        return [TextContent(
            type="text",
            text="Errore: specificare pipeline_id"
        )]

    client = CircleCIClient()

    if not client.is_available():
        return [TextContent(
            type="text",
            text="Errore: CircleCI non configurato. Assicurarsi che CIRCLECI_TOKEN sia impostato."
        )]

    # Get pipeline info
    pipeline = client.get_pipeline(pipeline_id)
    if not pipeline:
        return [TextContent(
            type="text",
            text=f"Pipeline {pipeline_id} non trovata."
        )]

    # Get workflows
    workflows = client.get_pipeline_workflows(pipeline_id)

    result = f"""**Pipeline #{pipeline.number}** {pipeline.status_icon}

- Stato: {pipeline.state}
- Creata: {pipeline.created_at[:19].replace('T', ' ') if pipeline.created_at else '?'}
- URL: {pipeline.url}

"""

    if workflows:
        result += "**Workflows:**\n\n"
        for w in workflows:
            result += f"{w.status_icon} **{w.name}** [{w.status}]\n"

            # Get jobs for workflow
            jobs = client.get_workflow_jobs(w.id)
            if jobs:
                for job in jobs:
                    job_status = job.get("status", "unknown")
                    job_name = job.get("name", "?")
                    job_icon = {
                        "running": "...",
                        "success": "ok",
                        "failed": "X",
                        "blocked": "-",
                        "canceled": "X",
                        "not_run": "-"
                    }.get(job_status, "?")
                    result += f"   [{job_icon}] {job_name}\n"
            result += "\n"
    else:
        result += "*Nessun workflow ancora avviato*\n"

    return [TextContent(type="text", text=result)]


def get_sheets_client(project: str):
    """Get a GoogleSheetsClient for a project."""
    from src.sheets_client import GoogleSheetsClient
    from src.config_loader import ConfigLoader
    from pathlib import Path

    logger.info(f"get_sheets_client called for project: {project}")

    # Load project configuration
    project_root = Path(__file__).parent.parent
    logger.info(f"Project root: {project_root}")

    loader = ConfigLoader(str(project_root))
    logger.info(f"ConfigLoader created")

    project_config = loader.load_project(project)
    logger.info(f"Project config loaded: {project_config.name}")

    # Get Google Sheets config
    gs_config = project_config.google_sheets
    logger.info(f"Google Sheets enabled: {gs_config.enabled}, spreadsheet_id: {gs_config.spreadsheet_id}")

    if not gs_config.enabled:
        raise Exception(f"Google Sheets not enabled for project {project}")

    # Resolve credentials path relative to project root
    credentials_path = gs_config.credentials_path or "config/oauth_credentials.json"
    full_credentials_path = project_root / credentials_path
    logger.info(f"Credentials path: {full_credentials_path}, exists: {full_credentials_path.exists()}")

    client = GoogleSheetsClient(
        credentials_path=str(full_credentials_path),
        spreadsheet_id=gs_config.spreadsheet_id,
        drive_folder_id=gs_config.drive_folder_id
    )
    logger.info("GoogleSheetsClient created, authenticating...")

    # Authenticate
    if not client.authenticate():
        raise Exception("Failed to authenticate with Google Sheets")

    logger.info("Authentication successful")
    return client


async def handle_list_runs(arguments: dict) -> list[TextContent]:
    """Handle list_runs tool."""
    project = arguments.get("project")

    if not project:
        return [TextContent(type="text", text="Errore: specificare il nome del progetto")]

    if project not in get_available_projects():
        return [TextContent(
            type="text",
            text=f"Progetto '{project}' non trovato. Usa list_projects per vedere i progetti disponibili."
        )]

    try:
        client = get_sheets_client(project)
        runs = client.get_all_run_numbers()

        if not runs:
            return [TextContent(type="text", text=f"Nessuna RUN trovata per il progetto '{project}'.")]

        result = f"**RUN disponibili per {project}:**\n\n"

        # Get info for last 10 runs (most recent first)
        for run_num in sorted(runs, reverse=True)[:10]:
            info = client.get_run_info(run_num)
            if info:
                timestamp = info.get("timestamp", "?")
                env = info.get("env", "?")
                mode = info.get("mode", "?")
                total = info.get("total_tests", 0)
                passed = info.get("passed", 0)
                failed = info.get("failed", 0)
                pass_rate = (passed / total * 100) if total > 0 else 0
                result += f"- **RUN {run_num}** ({timestamp}) - {env}/{mode}\n"
                result += f"  {passed}/{total} passed ({pass_rate:.0f}%), {failed} failed\n"
            else:
                result += f"- **RUN {run_num}**\n"

        if len(runs) > 10:
            result += f"\n... e altre {len(runs) - 10} RUN precedenti"

        return [TextContent(type="text", text=result)]
    except Exception as e:
        logger.error(f"Error listing runs: {e}")
        return [TextContent(type="text", text=f"Errore nel recupero delle RUN: {str(e)}")]


async def handle_get_run_results(arguments: dict) -> list[TextContent]:
    """Handle get_run_results tool."""
    project = arguments.get("project")
    run_number = arguments.get("run_number")

    if not project:
        return [TextContent(type="text", text="Errore: specificare il nome del progetto")]

    if project not in get_available_projects():
        return [TextContent(
            type="text",
            text=f"Progetto '{project}' non trovato."
        )]

    try:
        client = get_sheets_client(project)

        # Get latest run if not specified
        if not run_number:
            runs = client.get_all_run_numbers()
            if not runs:
                return [TextContent(type="text", text="Nessuna RUN trovata.")]
            run_number = max(runs)

        # Load the worksheet for this run
        worksheet = client.get_run_sheet(run_number)
        if not worksheet:
            return [TextContent(type="text", text=f"RUN {run_number} non trovata nel foglio Google Sheets.")]

        sheet_name = worksheet.title
        client._worksheet = worksheet
        results = client.get_all_results()

        if not results:
            return [TextContent(type="text", text=f"Nessun dato trovato per RUN {run_number} (foglio: {sheet_name}).")]

        # Calculate stats (ESITO è opzionale, compilato manualmente dal reviewer)
        total = len(results)
        reviewed = sum(1 for r in results if get_field(r, "esito").strip())
        passed = sum(1 for r in results if get_field(r, "esito").upper() == "PASS")
        failed = sum(1 for r in results if get_field(r, "esito").upper() == "FAIL")

        result = f"""**RUN {run_number}** - {project}
📄 Foglio: `{sheet_name}`
📊 Test registrati: **{total}**
"""
        # Mostra statistiche ESITO solo se ci sono review
        if reviewed > 0:
            pass_rate = (passed / reviewed * 100) if reviewed > 0 else 0
            result += f"""
**Review completate:** {reviewed}/{total}
- ✅ Passati: {passed} ({pass_rate:.1f}%)
- ❌ Falliti: {failed}
- ⏳ Da revieware: {total - reviewed}
"""
        else:
            result += "\n⏳ **Nessuna review completata** (colonna ESITO vuota)\n"

        result += "\n**Dati registrati:**\n\n"

        # Mostra i dati effettivi del foglio
        for r in results[:20]:  # Limit to 20 results
            test_id = get_field(r, "test_id", "?")
            date = get_field(r, "date", "")
            question = get_field(r, "question", "")[:60]
            conversation = get_field(r, "conversation", "")
            esito = get_field(r, "esito", "").upper()

            # Icona basata su: ha conversazione? ha esito?
            if esito == "PASS":
                icon = "✅"
            elif esito == "FAIL":
                icon = "❌"
            elif conversation:
                icon = "📝"  # Ha dati ma non reviewato
            else:
                icon = "⏳"  # Vuoto

            result += f"{icon} `{test_id}`"
            if date:
                # Estrai solo ora se è oggi
                date_short = date.split(" ")[-1] if " " in date else date
                result += f" | {date_short}"
            if question:
                result += f" | {question}..."
            result += "\n"

        if len(results) > 20:
            result += f"\n... e altri {len(results) - 20} test"

        return [TextContent(type="text", text=result)]
    except Exception as e:
        logger.error(f"Error getting run results: {e}")
        return [TextContent(type="text", text=f"Errore nel recupero dei risultati: {str(e)}")]


async def handle_get_failed_tests(arguments: dict) -> list[TextContent]:
    """Handle get_failed_tests tool."""
    project = arguments.get("project")
    run_number = arguments.get("run_number")

    if not project:
        return [TextContent(type="text", text="Errore: specificare il nome del progetto")]

    if project not in get_available_projects():
        return [TextContent(
            type="text",
            text=f"Progetto '{project}' non trovato."
        )]

    try:
        client = get_sheets_client(project)

        # Get latest run if not specified
        if not run_number:
            runs = client.get_all_run_numbers()
            if not runs:
                return [TextContent(type="text", text="Nessuna RUN trovata.")]
            run_number = max(runs)

        # Load the worksheet for this run
        worksheet = client.get_run_sheet(run_number)
        if not worksheet:
            return [TextContent(type="text", text=f"RUN {run_number} non trovata nel foglio Google Sheets.")]

        client._worksheet = worksheet
        results = client.get_all_results()

        # Filter failed tests
        failed_tests = [r for r in results if get_field(r, "esito").upper() == "FAIL"]

        if not failed_tests:
            return [TextContent(type="text", text=f"Nessun test fallito nella RUN {run_number}!")]

        result = f"""**Test Falliti - RUN {run_number}** ({len(failed_tests)} fallimenti)

"""
        for r in failed_tests:
            test_id = get_field(r, "test_id", "?")
            question = r.get("question", r.get("domanda", ""))[:80]
            notes = r.get("notes", r.get("note", ""))[:100]
            result += f"**{test_id}**\n"
            if question:
                result += f"  Domanda: {question}...\n"
            if notes:
                result += f"  Note: {notes}...\n"
            result += "\n"

        return [TextContent(type="text", text=result)]
    except Exception as e:
        logger.error(f"Error getting failed tests: {e}")
        return [TextContent(type="text", text=f"Errore nel recupero dei test falliti: {str(e)}")]


async def handle_compare_runs(arguments: dict) -> list[TextContent]:
    """Handle compare_runs tool."""
    from src.comparison import RunComparator

    project = arguments.get("project")
    run_a = arguments.get("run_a")
    run_b = arguments.get("run_b")

    if not project:
        return [TextContent(type="text", text="Errore: specificare il nome del progetto")]

    if project not in get_available_projects():
        return [TextContent(
            type="text",
            text=f"Progetto '{project}' non trovato."
        )]

    try:
        comparator = RunComparator(project)

        # If no runs specified, compare latest two
        if not run_a and not run_b:
            comparison = comparator.compare_latest()
            if not comparison:
                return [TextContent(type="text", text="Servono almeno 2 RUN per il confronto.")]
        elif run_a and run_b:
            comparison = comparator.compare(run_a, run_b)
        else:
            # Get latest run for comparison
            client = get_sheets_client(project)
            runs = client.get_all_run_numbers()
            if not runs:
                return [TextContent(type="text", text="Nessuna RUN trovata.")]
            if run_a:
                run_b = max(runs)
            else:
                run_a = sorted(runs)[-2] if len(runs) > 1 else runs[0]
                run_b = max(runs)
            comparison = comparator.compare(run_a, run_b)

        # Format results
        result = f"""**Confronto RUN {comparison.run_a} vs RUN {comparison.run_b}**

**Riepilogo:**
- RUN {comparison.run_a}: {comparison.passed_a}/{comparison.total_tests_a} passed ({comparison.pass_rate_a:.1f}%)
- RUN {comparison.run_b}: {comparison.passed_b}/{comparison.total_tests_b} passed ({comparison.pass_rate_b:.1f}%)
- Delta: {comparison.pass_rate_delta:+.1f}%

"""
        if comparison.regressions:
            result += f"**Regressioni ({len(comparison.regressions)}):** (erano PASS, ora FAIL)\n"
            for r in comparison.regressions[:10]:
                result += f"- `{r.test_id}`: {r.old_result} -> {r.new_result}\n"
            if len(comparison.regressions) > 10:
                result += f"  ... e altre {len(comparison.regressions) - 10}\n"
            result += "\n"

        if comparison.improvements:
            result += f"**Miglioramenti ({len(comparison.improvements)}):** (erano FAIL, ora PASS)\n"
            for r in comparison.improvements[:10]:
                result += f"- `{r.test_id}`: {r.old_result} -> {r.new_result}\n"
            if len(comparison.improvements) > 10:
                result += f"  ... e altri {len(comparison.improvements) - 10}\n"
            result += "\n"

        if not comparison.regressions and not comparison.improvements:
            result += "*Nessun cambiamento significativo tra le due RUN.*\n"

        return [TextContent(type="text", text=result)]
    except Exception as e:
        logger.error(f"Error comparing runs: {e}")
        return [TextContent(type="text", text=f"Errore nel confronto delle RUN: {str(e)}")]


# ====== NUOVI HANDLER FOOL-PROOF ======

async def handle_start_test_session(arguments: dict) -> list[TextContent]:
    """Handle start_test_session tool - wizard guidato per testare un progetto."""
    project = arguments.get("project")

    if not project:
        return [TextContent(type="text", text="Errore: specificare il nome del progetto")]

    if project not in get_available_projects():
        projects = get_available_projects()
        return [TextContent(
            type="text",
            text=f"Progetto '{project}' non trovato.\n\nProgetti disponibili: {', '.join(projects)}"
        )]

    # Get test sets with counts
    test_sets = get_available_test_sets(project)
    run_config = get_run_config(project)

    # Try to get current run info from Google Sheets
    current_run = run_config.get("active_run", 0)
    env = run_config.get("env", "DEV")
    prompt_version = run_config.get("prompt_version", "N/A")

    result = f"Ecco la situazione per **{project}**:\n\n"

    # For each test set, show details
    for name, total_count in sorted(test_sets.items()):
        result += f"📋 **{name}** ({total_count} test)\n"

        # Try to get executed count from Sheets
        try:
            if current_run > 0:
                client = get_sheets_client(project)
                client.active_run = current_run
                results = client.get_all_results()

                # Filter results for this test set
                prefix = "PARA_" if name == "paraphrase" else "GRD_" if name == "ggp" else "TEST_"
                set_results = [r for r in results if r.get("test_id", "").startswith(prefix)]

                if set_results:
                    passed = sum(1 for r in set_results if get_field(r, "esito").upper() == "PASS")
                    failed = sum(1 for r in set_results if get_field(r, "esito").upper() == "FAIL")
                    pending = total_count - len(set_results)
                    pass_rate = (passed / len(set_results) * 100) if set_results else 0
                    result += f"   - {pending} pending, {failed} falliti, pass rate {pass_rate:.0f}%\n"
                else:
                    result += f"   - Tutti pending\n"
            else:
                result += f"   - Tutti pending (nessuna RUN attiva)\n"
        except Exception as e:
            logger.warning(f"Could not get stats for {project}/{name}: {e}")
            result += f"   - Stato non disponibile\n"

        result += "\n"

    # Add run info
    if current_run > 0:
        result += f"📈 **RUN attiva:** #{current_run}\n"
        result += f"📝 **Prompt:** {prompt_version}\n"
        result += f"🌍 **Ambiente:** {env}\n\n"

    # Add suggestion
    result += "---\n\n"
    if test_sets.get("standard", 0) > 0:
        result += "💡 **Consiglio:** eseguire i test pending del set **standard**.\n\n"
    result += "**Quale test set vuoi usare?** (standard, paraphrase, ggp)"

    return [TextContent(type="text", text=result)]


async def handle_prepare_test_run(arguments: dict) -> list[TextContent]:
    """Handle prepare_test_run tool - prepara e mostra riepilogo per conferma."""
    project = arguments.get("project")
    test_set = arguments.get("test_set", "standard")
    tests_filter = arguments.get("tests", "pending")

    if not project:
        return [TextContent(type="text", text="Errore: specificare il nome del progetto")]

    if project not in get_available_projects():
        return [TextContent(type="text", text=f"Progetto '{project}' non trovato.")]

    # Load configs
    run_config = get_run_config(project)
    all_tests = get_project_tests(project, test_set)

    if not all_tests:
        return [TextContent(type="text", text=f"Nessun test trovato per il set '{test_set}'.")]

    # Count tests to run
    test_count = len(all_tests)
    executed_ids = set()
    failed_ids = set()

    try:
        current_run = run_config.get("active_run", 0)
        if current_run > 0:
            client = get_sheets_client(project)
            client.active_run = current_run
            results = client.get_all_results()
            executed_ids = {r.get("test_id") for r in results if r.get("test_id")}
            failed_ids = {r.get("test_id") for r in results
                         if get_field(r, "esito").upper() == "FAIL"}
    except Exception as e:
        logger.warning(f"Could not get executed tests: {e}")

    # Filter tests based on mode
    if tests_filter == "pending":
        tests_to_run = [t for t in all_tests if t.get("id") not in executed_ids]
        test_count = len(tests_to_run)
    elif tests_filter == "failed":
        tests_to_run = [t for t in all_tests if t.get("id") in failed_ids]
        test_count = len(tests_to_run)
    else:  # all
        test_count = len(all_tests)

    if test_count == 0:
        return [TextContent(
            type="text",
            text=f"Nessun test da eseguire con filtro '{tests_filter}' per il set '{test_set}'."
        )]

    # Determine if new RUN
    current_run = run_config.get("active_run", 0)
    new_run = (tests_filter == "all") or (current_run == 0)
    next_run = current_run + 1 if new_run else current_run

    # Auto-determine parallelism
    use_parallel = should_use_parallel(test_count)
    estimated_time = estimate_duration(test_count, use_parallel)

    # Sheet name
    sheet_name = f"RUN_{next_run}"

    # Build summary
    env = run_config.get("env", "DEV")
    prompt_version = run_config.get("prompt_version", "N/A")

    result = f"""📋 **RIEPILOGO ESECUZIONE**

🎯 **Progetto:** {project}
🌍 **Ambiente:** {env}
📝 **Prompt version:** {prompt_version}

📊 **Test Set:** {test_set}
   - {test_count} test da eseguire ({tests_filter})
   - Modalità: auto

📈 **RUN:** #{next_run} {'(NUOVA)' if new_run else '(ESISTENTE)'}
📄 **Google Sheets:** Foglio "{sheet_name}"

⚡ **Parallelismo:** {'SÌ (3 container)' if use_parallel else 'NO (singolo container)'}
⏱️ **Tempo stimato:** ~{estimated_time}

---
**Procedo con l'esecuzione?** (sì/no)"""

    return [TextContent(type="text", text=result)]


async def handle_execute_test_run(arguments: dict) -> list[TextContent]:
    """Handle execute_test_run tool - esegue dopo conferma."""
    from src.circleci_client import CircleCIClient

    project = arguments.get("project")
    test_set = arguments.get("test_set", "standard")
    tests_filter = arguments.get("tests", "pending")

    if not project:
        return [TextContent(type="text", text="Errore: specificare il nome del progetto")]

    if project not in get_available_projects():
        return [TextContent(type="text", text=f"Progetto '{project}' non trovato.")]

    # Load configs
    run_config = get_run_config(project)
    all_tests = get_project_tests(project, test_set)

    # Count tests to determine parallelism
    test_count = len(all_tests)

    try:
        current_run = run_config.get("active_run", 0)
        if current_run > 0 and tests_filter in ["pending", "failed"]:
            client = get_sheets_client(project)
            client.active_run = current_run
            results = client.get_all_results()
            executed_ids = {r.get("test_id") for r in results if r.get("test_id")}
            failed_ids = {r.get("test_id") for r in results
                         if get_field(r, "esito").upper() == "FAIL"}

            if tests_filter == "pending":
                test_count = len([t for t in all_tests if t.get("id") not in executed_ids])
            elif tests_filter == "failed":
                test_count = len([t for t in all_tests if t.get("id") in failed_ids])
    except Exception as e:
        logger.warning(f"Could not count tests: {e}")

    # Auto-determine parallelism
    use_parallel = should_use_parallel(test_count)
    native_parallelism = 3 if use_parallel else 0

    # Determine new_run
    new_run = (tests_filter == "all")

    # Build tests_file
    tests_file = "tests.json" if test_set == "standard" else f"tests_{test_set}.json"

    # Trigger CircleCI
    circleci = CircleCIClient()

    if not circleci.is_available():
        return [TextContent(
            type="text",
            text="❌ Errore: CircleCI non configurato. Assicurarsi che CIRCLECI_TOKEN sia impostato."
        )]

    success, response = circleci.trigger_pipeline(
        project=project,
        mode="auto",
        tests=tests_filter,
        new_run=new_run,
        tests_file=tests_file,
        native_parallelism=native_parallelism
    )

    if not success:
        error = response.get("error", "Errore sconosciuto") if response else "Errore sconosciuto"
        return [TextContent(type="text", text=f"❌ Errore nell'avvio della pipeline: {error}")]

    pipeline_id = response.get("id", "?")
    pipeline_number = response.get("number", "?")
    pipeline_url = f"https://app.circleci.com/pipelines/gh/corradofrancolini/chatbot-tester-private/{pipeline_number}"

    result = f"""✅ **Pipeline avviata!**

🔗 **URL:** {pipeline_url}
🆔 **ID:** `{pipeline_id}`

{'⚡ Parallelismo attivo (3 container)' if use_parallel else ''}

⏰ **Chiedimi "come sta andando?" tra 5-10 minuti** per un aggiornamento."""

    return [TextContent(type="text", text=result)]


async def handle_check_pipeline_status(arguments: dict) -> list[TextContent]:
    """Handle check_pipeline_status tool - controlla stato pipeline."""
    from src.circleci_client import CircleCIClient

    pipeline_id = arguments.get("pipeline_id")

    if not pipeline_id:
        return [TextContent(type="text", text="Errore: specificare pipeline_id")]

    client = CircleCIClient()

    if not client.is_available():
        return [TextContent(
            type="text",
            text="❌ Errore: CircleCI non configurato."
        )]

    # Get pipeline info
    pipeline = client.get_pipeline(pipeline_id)
    if not pipeline:
        return [TextContent(type="text", text=f"Pipeline `{pipeline_id}` non trovata.")]

    # Get workflows
    workflows = client.get_pipeline_workflows(pipeline_id)

    # Check if all done
    all_done = all(w.status in ["success", "failed", "canceled"]
                   for w in workflows) if workflows else False

    if all_done:
        success = all(w.status == "success" for w in workflows)
        icon = "✅" if success else "❌"
        status = "SUCCESS" if success else "FAILED"

        return [TextContent(
            type="text",
            text=f"""{icon} **Pipeline completata!**

Stato: **{status}**
URL: {pipeline.url}

Vuoi vedere i risultati dei test? Chiedimi "mostrami i risultati di [progetto]"."""
        )]

    # Still running
    jobs_info = ""
    if workflows:
        try:
            jobs = client.get_workflow_jobs(workflows[0].id)
            running = [j for j in jobs if j.get("status") == "running"]
            if running:
                jobs_info = f"\nJob in esecuzione: **{running[0].get('name', '?')}**"
        except Exception:
            pass

    return [TextContent(
        type="text",
        text=f"""⏳ **Pipeline in esecuzione...**

Stato: **{workflows[0].status if workflows else 'starting'}**{jobs_info}
URL: {pipeline.url}

⏰ **Chiedimi di nuovo tra 3-5 minuti** per un aggiornamento."""
    )]


async def handle_show_results(arguments: dict) -> list[TextContent]:
    """Handle show_results tool - mostra risultati user-friendly."""
    project = arguments.get("project")
    run_number = arguments.get("run_number")

    if not project:
        return [TextContent(type="text", text="Errore: specificare il nome del progetto")]

    if project not in get_available_projects():
        return [TextContent(type="text", text=f"Progetto '{project}' non trovato.")]

    try:
        client = get_sheets_client(project)

        # Get latest run if not specified
        if not run_number:
            runs = client.get_all_run_numbers()
            if not runs:
                return [TextContent(type="text", text="Nessuna RUN trovata.")]
            run_number = max(runs)

        # Load the worksheet for this run
        worksheet = client.get_run_sheet(run_number)
        if not worksheet:
            return [TextContent(type="text", text=f"RUN {run_number} non trovata nel foglio Google Sheets.")]

        client._worksheet = worksheet
        results = client.get_all_results()

        if not results:
            return [TextContent(type="text", text=f"Nessun risultato trovato per RUN {run_number}.")]

        # Calculate stats
        total = len(results)
        passed = sum(1 for r in results if get_field(r, "esito").upper() == "PASS")
        failed = sum(1 for r in results if get_field(r, "esito").upper() == "FAIL")
        pass_rate = (passed / total * 100) if total > 0 else 0

        # Choose icon based on pass rate
        if pass_rate >= 90:
            icon = "✅"
        elif pass_rate >= 70:
            icon = "⚠️"
        else:
            icon = "❌"

        result = f"""{icon} **Risultati RUN #{run_number} - {project}**

**Pass rate: {pass_rate:.0f}%** ({passed}/{total})

"""
        # Show failed tests
        failed_tests = [r for r in results if get_field(r, "esito").upper() == "FAIL"]

        if failed_tests:
            result += f"❌ **{len(failed_tests)} test falliti:**\n"
            for r in failed_tests[:10]:
                test_id = get_field(r, "test_id", "?")
                question = r.get("question", r.get("domanda", ""))[:50]
                result += f"- `{test_id}`: {question}...\n"
            if len(failed_tests) > 10:
                result += f"  ... e altri {len(failed_tests) - 10}\n"
            result += "\n"

        # Add suggestions
        result += "---\n"
        if failed_tests:
            result += f"💡 Vuoi ri-eseguire solo i {len(failed_tests)} test falliti? Dimmelo!"
        else:
            result += "🎉 Tutti i test sono passati!"

        return [TextContent(type="text", text=result)]

    except Exception as e:
        logger.error(f"Error showing results: {e}")
        return [TextContent(type="text", text=f"Errore nel recupero dei risultati: {str(e)}")]


async def handle_add_test(arguments: dict) -> list[TextContent]:
    """Handle add_test tool - adds a new test to a test set."""
    project = arguments.get("project")
    test_set = arguments.get("test_set", "standard")
    question = arguments.get("question")
    expected_answer = arguments.get("expected_answer")
    category = arguments.get("category")
    follow_ups = arguments.get("follow_ups", [])

    if not project:
        return [TextContent(type="text", text="Errore: specificare il nome del progetto")]

    if not question:
        return [TextContent(type="text", text="Errore: specificare la domanda del test")]

    projects = get_available_projects()
    if project not in projects:
        return [TextContent(
            type="text",
            text=f"Progetto '{project}' non trovato.\n\nProgetti disponibili: {', '.join(projects)}"
        )]

    # Determine tests file path
    project_dir = PROJECTS_DIR / project
    if test_set == "standard":
        tests_file = project_dir / "tests.json"
    else:
        tests_file = project_dir / f"tests_{test_set}.json"

    # Load existing tests
    existing_tests = []
    if tests_file.exists():
        try:
            with open(tests_file) as f:
                data = json.load(f)
                if isinstance(data, list):
                    existing_tests = data
                elif isinstance(data, dict) and "tests" in data:
                    existing_tests = data["tests"]
        except Exception as e:
            logger.error(f"Error reading tests file: {e}")
            return [TextContent(type="text", text=f"Errore nel leggere il file test: {str(e)}")]

    # Generate new test ID
    # Determine prefix based on test set
    if test_set == "paraphrase":
        prefix = "PARA"
    elif test_set == "ggp":
        prefix = "GGP"
    else:
        prefix = "TEST"

    # Find max existing ID with this prefix
    max_num = 0
    for test in existing_tests:
        test_id = test.get("id", "")
        if test_id.startswith(prefix + "_"):
            try:
                num = int(test_id.split("_")[1])
                max_num = max(max_num, num)
            except (ValueError, IndexError):
                pass

    new_id = f"{prefix}_{max_num + 1:03d}"

    # Build new test object
    new_test = {
        "id": new_id,
        "question": question
    }

    if expected_answer:
        new_test["expected_answer"] = expected_answer

    if category:
        new_test["category"] = category

    if follow_ups:
        new_test["follow_ups"] = follow_ups

    # Add to tests list
    existing_tests.append(new_test)

    # Write back to file
    try:
        with open(tests_file, "w", encoding="utf-8") as f:
            json.dump(existing_tests, f, indent=2, ensure_ascii=False)

        result = f"""✅ **Test aggiunto con successo!**

**ID:** `{new_id}`
**Progetto:** {project}
**Test set:** {test_set}
**File:** {tests_file.name}

**Domanda:** {question}
"""
        if expected_answer:
            result += f"**Risposta attesa:** {expected_answer}\n"
        if category:
            result += f"**Categoria:** {category}\n"
        if follow_ups:
            result += f"**Follow-ups:** {len(follow_ups)} domande\n"

        result += f"\n📊 Totale test nel set: {len(existing_tests)}"
        result += f"\n\n💡 Per eseguire questo test: *\"Esegui {new_id} su {project}\"*"

        return [TextContent(type="text", text=result)]

    except Exception as e:
        logger.error(f"Error writing tests file: {e}")
        return [TextContent(type="text", text=f"Errore nel salvare il test: {str(e)}")]


async def handle_notify_corrado(arguments: dict) -> list[TextContent]:
    """Handle notify_corrado tool - sends a Telegram notification."""
    import httpx
    from datetime import datetime

    message = arguments.get("message")
    project = arguments.get("project")
    priority = arguments.get("priority", "normal")

    if not message:
        return [TextContent(type="text", text="Errore: specificare il messaggio da inviare")]

    # Get Telegram credentials from environment
    bot_token = os.environ.get("TELEGRAM_BOT_TOKEN")
    chat_id = os.environ.get("TELEGRAM_CHAT_ID")

    if not bot_token or not chat_id:
        logger.error("Telegram credentials not configured")
        return [TextContent(
            type="text",
            text="Errore: Telegram non configurato. Contatta l'amministratore."
        )]

    # Build the notification message
    priority_emoji = {"low": "📝", "normal": "🔔", "high": "🚨"}
    emoji = priority_emoji.get(priority, "🔔")

    timestamp = datetime.now().strftime("%d/%m/%Y %H:%M")

    telegram_message = f"""{emoji} *Notifica da Claude Desktop*

📋 *Progetto:* {project or "Non specificato"}
💬 *Messaggio:* {message}

🕐 {timestamp}"""

    # Send via Telegram API
    telegram_url = f"https://api.telegram.org/bot{bot_token}/sendMessage"

    try:
        async with httpx.AsyncClient() as client:
            response = await client.post(
                telegram_url,
                json={
                    "chat_id": chat_id,
                    "text": telegram_message,
                    "parse_mode": "Markdown"
                },
                timeout=10.0
            )

            if response.status_code == 200:
                return [TextContent(
                    type="text",
                    text=f"✅ Notifica inviata a Corrado!\n\nMessaggio: {message}"
                )]
            else:
                logger.error(f"Telegram API error: {response.status_code} - {response.text}")
                return [TextContent(
                    type="text",
                    text=f"Errore nell'invio della notifica: {response.status_code}"
                )]

    except Exception as e:
        logger.error(f"Error sending Telegram notification: {e}")
        return [TextContent(type="text", text=f"Errore nell'invio della notifica: {str(e)}")]


async def handle_novita(arguments: dict) -> list[TextContent]:
    """Handle novita tool - shows changelog."""
    limit = arguments.get("limit", 3)

    if not CHANGELOG:
        return [TextContent(type="text", text="Nessuna novità disponibile.")]

    result = "# 🆕 Novità del Sistema di Testing\n\n"

    for i, release in enumerate(CHANGELOG[:limit]):
        version = release.get("version", "?")
        date = release.get("date", "?")
        title = release.get("title", "")
        changes = release.get("changes", [])

        # Header versione
        if i == 0:
            result += f"## ✨ Versione {version} — {title}\n"
            result += f"📅 Rilasciata il {date}\n\n"
        else:
            result += f"---\n\n## Versione {version} — {title}\n"
            result += f"📅 {date}\n\n"

        # Cambiamenti con formato dettagliato
        for change in changes:
            if isinstance(change, dict):
                icon = change.get("icon", "•")
                change_title = change.get("title", "")
                description = change.get("description", "")
                result += f"### {icon} {change_title}\n"
                result += f"{description}\n\n"
            else:
                # Fallback per formato vecchio (stringa semplice)
                result += f"- {change}\n"

        result += "\n"

    if len(CHANGELOG) > limit:
        result += f"---\n*Altre {len(CHANGELOG) - limit} versioni precedenti disponibili*"

    return [TextContent(type="text", text=result)]


# ====== HANDLER NUOVI TOOL v1.4.0 ======

async def handle_detect_flaky_tests(arguments: dict) -> list[TextContent]:
    """Handle detect_flaky_tests tool."""
    from src.comparison import RunComparator, FlakyTestDetector

    project = arguments.get("project")
    last_n_runs = arguments.get("last_n_runs", 10)
    threshold = arguments.get("threshold", 0.3)

    if not project:
        return [TextContent(type="text", text="Errore: specificare il nome del progetto")]

    if project not in get_available_projects():
        return [TextContent(type="text", text=f"Progetto '{project}' non trovato.")]

    try:
        client = get_sheets_client(project)
        comparator = RunComparator(client)
        detector = FlakyTestDetector(comparator)

        flaky_tests = detector.detect_flaky_tests(
            last_n_runs=last_n_runs,
            flaky_threshold=threshold
        )

        if not flaky_tests:
            return [TextContent(type="text", text=f"✅ Nessun test flaky rilevato nelle ultime {last_n_runs} run di **{project}**")]

        result = f"**Test Flaky in {project}** (ultime {last_n_runs} run)\n\n"
        result += f"Soglia: {threshold} (0=stabile, 1=random)\n\n"

        for ft in flaky_tests[:15]:
            score_bar = "🔴" if ft.flaky_score > 0.6 else "🟡" if ft.flaky_score > 0.3 else "🟢"
            result += f"{score_bar} **{ft.test_id}** - score: {ft.flaky_score:.2f}\n"
            result += f"   PASS: {ft.pass_count}/{ft.total_runs} | FAIL: {ft.fail_count}/{ft.total_runs}\n"

        if len(flaky_tests) > 15:
            result += f"\n... e altri {len(flaky_tests) - 15} test flaky"

        return [TextContent(type="text", text=result)]
    except Exception as e:
        logger.error(f"Error detecting flaky tests: {e}")
        return [TextContent(type="text", text=f"Errore nel rilevamento flaky tests: {str(e)}")]


async def handle_get_regressions(arguments: dict) -> list[TextContent]:
    """Handle get_regressions tool."""
    from src.comparison import RunComparator, RegressionDetector

    project = arguments.get("project")
    run_number = arguments.get("run_number")

    if not project:
        return [TextContent(type="text", text="Errore: specificare il nome del progetto")]

    if project not in get_available_projects():
        return [TextContent(type="text", text=f"Progetto '{project}' non trovato.")]

    try:
        client = get_sheets_client(project)

        # Get run number if not specified
        if not run_number:
            runs = client.get_all_run_numbers()
            if not runs:
                return [TextContent(type="text", text="Nessuna RUN trovata.")]
            run_number = max(runs)

        comparator = RunComparator(client)
        detector = RegressionDetector(comparator)

        regressions = detector.check_for_regressions(new_run=run_number)

        if not regressions:
            return [TextContent(type="text", text=f"✅ Nessuna regressione nella RUN {run_number} di **{project}**")]

        result = f"**Regressioni RUN {run_number}** - {project}\n\n"
        result += f"Test che erano PASS e ora sono FAIL:\n\n"

        for reg in regressions[:20]:
            result += f"❌ **{reg.test_id}**: {reg.old_result} → {reg.new_result}\n"
            if reg.category:
                result += f"   Categoria: {reg.category}\n"

        if len(regressions) > 20:
            result += f"\n... e altre {len(regressions) - 20} regressioni"

        return [TextContent(type="text", text=result)]
    except Exception as e:
        logger.error(f"Error getting regressions: {e}")
        return [TextContent(type="text", text=f"Errore nel rilevamento regressioni: {str(e)}")]


async def handle_get_performance_report(arguments: dict) -> list[TextContent]:
    """Handle get_performance_report tool."""
    from src.performance import PerformanceHistory, PerformanceReporter

    project = arguments.get("project")
    run_number = arguments.get("run_number")

    if not project:
        return [TextContent(type="text", text="Errore: specificare il nome del progetto")]

    if project not in get_available_projects():
        return [TextContent(type="text", text=f"Progetto '{project}' non trovato.")]

    try:
        reports_dir = Path(__file__).parent.parent / "reports"
        history = PerformanceHistory(project, reports_dir)

        # Load metrics
        metrics_list = history.load_history(last_n=5)

        if not metrics_list:
            return [TextContent(type="text", text=f"Nessun dato performance trovato per **{project}**. I dati vengono generati durante l'esecuzione dei test.")]

        # Get specific or latest
        if run_number:
            metrics = next((m for m in metrics_list if m.run_id == f"run_{run_number}"), None)
            if not metrics:
                return [TextContent(type="text", text=f"Dati performance non trovati per RUN {run_number}")]
        else:
            metrics = metrics_list[0]

        reporter = PerformanceReporter(metrics)
        summary = reporter.generate_summary()

        result = f"**Performance Report - {project}**\n\n"
        result += summary

        return [TextContent(type="text", text=result)]
    except Exception as e:
        logger.error(f"Error generating performance report: {e}")
        return [TextContent(type="text", text=f"Errore nel report performance: {str(e)}")]


async def handle_get_performance_alerts(arguments: dict) -> list[TextContent]:
    """Handle get_performance_alerts tool."""
    from src.performance import PerformanceHistory, PerformanceAlerter

    project = arguments.get("project")
    run_number = arguments.get("run_number")

    if not project:
        return [TextContent(type="text", text="Errore: specificare il nome del progetto")]

    if project not in get_available_projects():
        return [TextContent(type="text", text=f"Progetto '{project}' non trovato.")]

    try:
        reports_dir = Path(__file__).parent.parent / "reports"
        history = PerformanceHistory(project, reports_dir)

        metrics_list = history.load_history(last_n=5)

        if not metrics_list:
            return [TextContent(type="text", text=f"Nessun dato performance trovato per **{project}**")]

        current = metrics_list[0]
        baseline = metrics_list[1] if len(metrics_list) > 1 else None

        alerter = PerformanceAlerter()
        alerts = alerter.check(current, baseline)

        result = f"**Performance Alerts - {project}**\n\n"
        result += alerter.format_alerts()

        if alerter.has_critical_alerts():
            result += "\n\n🚨 **Attenzione:** Ci sono alert critici da investigare!"

        return [TextContent(type="text", text=result)]
    except Exception as e:
        logger.error(f"Error checking performance alerts: {e}")
        return [TextContent(type="text", text=f"Errore nel controllo alert: {str(e)}")]


async def handle_analyze_coverage(arguments: dict) -> list[TextContent]:
    """Handle analyze_coverage tool."""
    from src.comparison import CoverageAnalyzer

    project = arguments.get("project")
    test_set = arguments.get("test_set", "standard")

    if not project:
        return [TextContent(type="text", text="Errore: specificare il nome del progetto")]

    if project not in get_available_projects():
        return [TextContent(type="text", text=f"Progetto '{project}' non trovato.")]

    try:
        # Load tests from project directory
        project_dir = PROJECTS_DIR / project
        if test_set == "standard":
            tests_file = project_dir / "tests.json"
        else:
            tests_file = project_dir / f"tests_{test_set}.json"

        if not tests_file.exists():
            return [TextContent(type="text", text=f"File test non trovato: {tests_file.name}")]

        with open(tests_file) as f:
            tests = json.load(f)

        if not tests:
            return [TextContent(type="text", text=f"Nessun test trovato per {project}/{test_set}")]

        analyzer = CoverageAnalyzer()
        report = analyzer.analyze(tests)

        result = f"**Coverage Analysis - {project}/{test_set}**\n\n"
        result += f"Test totali: {report.total_tests}\n\n"

        result += "**Copertura per categoria:**\n"
        for category, count in sorted(report.categories.items(), key=lambda x: -x[1]):
            icon = "✅" if count >= 3 else "⚠️" if count >= 1 else "❌"
            result += f"{icon} {category}: {count} test\n"

        if report.uncovered_categories:
            result += f"\n**Categorie scoperte:**\n"
            for cat in report.uncovered_categories:
                result += f"❌ {cat}\n"

        if report.suggested_tests:
            result += f"\n**Suggerimenti:**\n"
            for suggestion in report.suggested_tests[:5]:
                result += f"💡 {suggestion}\n"

        return [TextContent(type="text", text=result)]
    except Exception as e:
        logger.error(f"Error analyzing coverage: {e}")
        return [TextContent(type="text", text=f"Errore nell'analisi coverage: {str(e)}")]


async def handle_get_stability_report(arguments: dict) -> list[TextContent]:
    """Handle get_stability_report tool."""
    from src.comparison import RunComparator, FlakyTestDetector

    project = arguments.get("project")
    last_n_runs = arguments.get("last_n_runs", 10)

    if not project:
        return [TextContent(type="text", text="Errore: specificare il nome del progetto")]

    if project not in get_available_projects():
        return [TextContent(type="text", text=f"Progetto '{project}' non trovato.")]

    try:
        client = get_sheets_client(project)
        comparator = RunComparator(client)
        detector = FlakyTestDetector(comparator)

        report = detector.get_stability_report(last_n_runs=last_n_runs)

        result = f"**Stability Report - {project}** (ultime {last_n_runs} run)\n\n"

        if isinstance(report, dict):
            result += f"📊 **Metriche:**\n"
            result += f"- RUN analizzate: {report.get('runs_analyzed', 0)}\n"
            result += f"- Test totali: {report.get('total_tests', 0)}\n"
            result += f"- Sempre PASS: {report.get('stable_pass', 0)} ✅\n"
            result += f"- Sempre FAIL: {report.get('stable_fail', 0)} ❌\n"
            result += f"- Flaky: {report.get('flaky_tests', 0)} ⚠️\n"
            result += f"\n**Stability Score:** {report.get('stability_score', 0):.1%}\n"

            if report.get('flaky_test_ids'):
                result += f"\n**Test flaky da investigare:**\n"
                for tid in report['flaky_test_ids'][:10]:
                    result += f"- {tid}\n"
        else:
            result += str(report)

        return [TextContent(type="text", text=result)]
    except Exception as e:
        logger.error(f"Error generating stability report: {e}")
        return [TextContent(type="text", text=f"Errore nel report stabilità: {str(e)}")]


async def handle_diagnose_prompt(arguments: dict) -> list[TextContent]:
    """Handle diagnose_prompt tool."""
    project = arguments.get("project")
    run_number = arguments.get("run_number")
    test_id = arguments.get("test_id")

    if not project:
        return [TextContent(type="text", text="Errore: specificare il nome del progetto")]

    if project not in get_available_projects():
        return [TextContent(type="text", text=f"Progetto '{project}' non trovato.")]

    try:
        client = get_sheets_client(project)

        if not run_number:
            runs = client.get_all_run_numbers()
            if not runs:
                return [TextContent(type="text", text="Nessuna RUN trovata.")]
            run_number = max(runs)

        worksheet = client.get_run_sheet(run_number)
        if not worksheet:
            return [TextContent(type="text", text=f"RUN {run_number} non trovata.")]

        client._worksheet = worksheet
        results = client.get_all_results()

        # Filter failed tests
        failed_tests = [r for r in results if get_field(r, "esito").upper() == "FAIL"]

        if test_id:
            failed_tests = [r for r in failed_tests if get_field(r, "test_id") == test_id]

        if not failed_tests:
            return [TextContent(type="text", text=f"Nessun test fallito trovato nella RUN {run_number}")]

        result = f"**Diagnosi Fallimenti - RUN {run_number}**\n\n"

        for ft in failed_tests[:5]:
            tid = get_field(ft, "test_id")
            question = get_field(ft, "question", "")[:100]
            expected = get_field(ft, "expected_answer", "")[:100]
            actual = get_field(ft, "conversation", "")[:200]
            notes = get_field(ft, "notes", "")

            result += f"### ❌ {tid}\n"
            result += f"**Domanda:** {question}\n"
            if expected:
                result += f"**Atteso:** {expected}\n"
            if actual:
                result += f"**Risposta:** {actual}...\n"

            # Simple heuristic diagnosis
            diagnosis = []
            if "italiano" in str(actual).lower() or "english" in str(question).lower():
                diagnosis.append("🔤 Possibile mismatch lingua")
            if len(str(actual)) < 20:
                diagnosis.append("📝 Risposta troppo breve")
            if "errore" in str(actual).lower() or "error" in str(actual).lower():
                diagnosis.append("⚠️ Errore nella risposta")
            if not diagnosis:
                diagnosis.append("🔍 Richiede analisi manuale")

            result += f"**Diagnosi:** {', '.join(diagnosis)}\n"
            if notes:
                result += f"**Note:** {notes}\n"
            result += "\n"

        if len(failed_tests) > 5:
            result += f"... e altri {len(failed_tests) - 5} test falliti"

        return [TextContent(type="text", text=result)]
    except Exception as e:
        logger.error(f"Error diagnosing prompt: {e}")
        return [TextContent(type="text", text=f"Errore nella diagnosi: {str(e)}")]


async def handle_calibrate_thresholds(arguments: dict) -> list[TextContent]:
    """Handle calibrate_thresholds tool."""
    project = arguments.get("project")
    last_n_runs = arguments.get("last_n_runs", 10)

    if not project:
        return [TextContent(type="text", text="Errore: specificare il nome del progetto")]

    if project not in get_available_projects():
        return [TextContent(type="text", text=f"Progetto '{project}' non trovato.")]

    try:
        client = get_sheets_client(project)
        runs = client.get_all_run_numbers()

        if not runs:
            return [TextContent(type="text", text="Nessuna RUN trovata.")]

        # Collect metrics from last N runs
        metrics_data = {"semantic": [], "judge": [], "overall": []}
        runs_analyzed = []

        for run_num in sorted(runs, reverse=True)[:last_n_runs]:
            worksheet = client.get_run_sheet(run_num)
            if worksheet:
                client._worksheet = worksheet
                results = client.get_all_results()
                runs_analyzed.append(run_num)

                for r in results:
                    sem = get_field(r, "semantic", "")
                    judge = get_field(r, "judge", "")
                    overall = get_field(r, "overall", "")

                    try:
                        if sem:
                            metrics_data["semantic"].append(float(sem))
                        if judge:
                            metrics_data["judge"].append(float(judge))
                        if overall:
                            metrics_data["overall"].append(float(overall))
                    except (ValueError, TypeError):
                        pass

        result = f"**Calibrazione Soglie - {project}**\n\n"
        result += f"RUN analizzate: {len(runs_analyzed)} ({min(runs_analyzed)}-{max(runs_analyzed)})\n\n"

        for metric_name, values in metrics_data.items():
            if values:
                import statistics
                result += f"**{metric_name.upper()}:**\n"
                result += f"  - Samples: {len(values)}\n"
                result += f"  - Min: {min(values):.2f}\n"
                result += f"  - Max: {max(values):.2f}\n"
                result += f"  - Media: {statistics.mean(values):.2f}\n"
                if len(values) > 1:
                    result += f"  - Mediana: {statistics.median(values):.2f}\n"
                    sorted_vals = sorted(values)
                    p25 = sorted_vals[len(sorted_vals)//4]
                    p75 = sorted_vals[3*len(sorted_vals)//4]
                    result += f"  - P25: {p25:.2f} | P75: {p75:.2f}\n"
                    # Suggested threshold: P25 for pass threshold
                    result += f"  - 💡 Soglia suggerita: {p25:.2f}\n"
                result += "\n"

        if not any(metrics_data.values()):
            result += "⚠️ Nessuna metrica trovata nei dati. Assicurati che le colonne SEMANTIC, JUDGE, OVERALL siano presenti nel foglio.\n"

        return [TextContent(type="text", text=result)]
    except Exception as e:
        logger.error(f"Error calibrating thresholds: {e}")
        return [TextContent(type="text", text=f"Errore nella calibrazione: {str(e)}")]


async def handle_export_report(arguments: dict) -> list[TextContent]:
    """Handle export_report tool."""
    from src.export import RunReport, ReportExporter

    project = arguments.get("project")
    run_number = arguments.get("run_number")
    export_format = arguments.get("format", "excel")

    if not project:
        return [TextContent(type="text", text="Errore: specificare il nome del progetto")]

    if project not in get_available_projects():
        return [TextContent(type="text", text=f"Progetto '{project}' non trovato.")]

    if export_format not in ["excel", "html", "csv"]:
        return [TextContent(type="text", text=f"Formato non supportato: {export_format}. Usa: excel, html, csv")]

    try:
        client = get_sheets_client(project)

        if not run_number:
            runs = client.get_all_run_numbers()
            if not runs:
                return [TextContent(type="text", text="Nessuna RUN trovata.")]
            run_number = max(runs)

        # Load data from sheets
        worksheet = client.get_run_sheet(run_number)
        if not worksheet:
            return [TextContent(type="text", text=f"RUN {run_number} non trovata.")]

        client._worksheet = worksheet
        results = client.get_all_results()

        # Create report dir
        reports_dir = Path(__file__).parent.parent / "reports" / project / f"run_{run_number}" / "exports"
        reports_dir.mkdir(parents=True, exist_ok=True)

        # Build simple report data
        total = len(results)
        passed = sum(1 for r in results if get_field(r, "esito").upper() == "PASS")
        failed = sum(1 for r in results if get_field(r, "esito").upper() == "FAIL")

        output_file = reports_dir / f"{project}_run{run_number}.{export_format if export_format != 'excel' else 'xlsx'}"

        if export_format == "csv":
            import csv
            import io
            # Generate CSV content in memory
            output = io.StringIO()
            if results:
                writer = csv.DictWriter(output, fieldnames=results[0].keys())
                writer.writeheader()
                writer.writerows(results)
            csv_content = output.getvalue()

            # Return content inline
            result = f"**📊 Export CSV - {project} RUN {run_number}**\n\n"
            result += f"Test: {total} | Pass: {passed} | Fail: {failed} | Rate: {passed/total*100:.1f}%\n\n"
            result += "```csv\n"
            result += csv_content
            result += "```"
            return [TextContent(type="text", text=result)]

        elif export_format == "html":
            html = f"""<!DOCTYPE html>
<html><head><meta charset="utf-8"><title>Report RUN {run_number}</title>
<style>body{{font-family:sans-serif;margin:20px}}table{{border-collapse:collapse;width:100%}}
th,td{{border:1px solid #ddd;padding:8px;text-align:left}}th{{background:#4CAF50;color:white}}
tr:nth-child(even){{background:#f9f9f9}}
.pass{{color:green;font-weight:bold}}.fail{{color:red;font-weight:bold}}</style></head>
<body><h1>📊 {project} - RUN {run_number}</h1>
<p><strong>Totale:</strong> {total} | <strong>Pass:</strong> {passed} | <strong>Fail:</strong> {failed} | <strong>Rate:</strong> {passed/total*100:.1f}%</p>
<table><tr><th>Test ID</th><th>Esito</th><th>Question</th><th>Response</th></tr>"""
            for r in results:
                esito = get_field(r, "esito").upper()
                css = "pass" if esito == "PASS" else "fail"
                question = get_field(r, 'question')[:100].replace('<', '&lt;').replace('>', '&gt;')
                response = get_field(r, 'response')[:150].replace('<', '&lt;').replace('>', '&gt;')
                html += f"<tr><td>{get_field(r, 'test_id')}</td><td class='{css}'>{esito}</td><td>{question}</td><td>{response}...</td></tr>"
            html += "</table></body></html>"

            # Return HTML content inline
            result = f"**📊 Export HTML - {project} RUN {run_number}**\n\n"
            result += f"Test: {total} | Pass: {passed} | Fail: {failed} | Rate: {passed/total*100:.1f}%\n\n"
            result += "```html\n"
            result += html
            result += "\n```\n\n"
            result += "_Copia il codice HTML e salvalo come file .html per visualizzarlo nel browser._"
            return [TextContent(type="text", text=result)]

        elif export_format == "excel":
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill

                wb = Workbook()
                ws = wb.active
                ws.title = "Results"

                # Headers
                headers = list(results[0].keys()) if results else []
                for col, h in enumerate(headers, 1):
                    ws.cell(row=1, column=col, value=h).font = Font(bold=True)

                # Data
                for row, r in enumerate(results, 2):
                    for col, h in enumerate(headers, 1):
                        ws.cell(row=row, column=col, value=r.get(h, ""))

                wb.save(output_file)

                # Excel can't be returned inline - suggest CSV
                result = f"**📊 Export Excel - {project} RUN {run_number}**\n\n"
                result += f"Test: {total} | Pass: {passed} | Fail: {failed} | Rate: {passed/total*100:.1f}%\n\n"
                result += f"⚠️ Il file Excel è stato generato sul server:\n`{output_file}`\n\n"
                result += "**Suggerimento:** Usa `format: csv` per ottenere i dati direttamente nella risposta."
                return [TextContent(type="text", text=result)]

            except ImportError:
                return [TextContent(type="text", text="openpyxl non installato. Usa formato csv o html.")]
    except Exception as e:
        logger.error(f"Error exporting report: {e}")
        return [TextContent(type="text", text=f"Errore nell'export: {str(e)}")]


async def handle_debug_trace(arguments: dict) -> list[TextContent]:
    """
    Fetch e analizza un trace LangSmith per debug.

    Mostra input/output, tool calls, timing e errori.
    """
    trace_id = arguments.get("trace_id", "").strip()
    project = arguments.get("project", "silicon-b")

    if not trace_id:
        return [TextContent(type="text", text="❌ Specifica un trace_id. Lo trovi nella colonna LS TRACE URL dei risultati.")]

    # Estrai solo l'ID se è un URL completo
    if "/" in trace_id:
        # URL format: https://smith.langchain.com/o/ORG/projects/p/PROJECT/r/TRACE_ID
        trace_id = trace_id.rstrip("/").split("/")[-1]
        # Rimuovi eventuali query params
        if "?" in trace_id:
            trace_id = trace_id.split("?")[0]

    # Carica config progetto per credenziali LangSmith
    project_dir = PROJECTS_DIR / project
    if not project_dir.exists():
        return [TextContent(type="text", text=f"❌ Progetto '{project}' non trovato")]

    try:
        from src.config_loader import ConfigLoader
        config_loader = ConfigLoader()
        proj_config = config_loader.load_project(project)

        if not proj_config:
            return [TextContent(type="text", text=f"❌ Impossibile caricare la configurazione di '{project}'")]

        if not proj_config.langsmith:
            return [TextContent(type="text", text=f"❌ LangSmith non configurato per '{project}'.\n\nConfigura langsmith in projects/{project}/project.yaml")]

        # Inizializza client
        from src.langsmith_client import LangSmithClient
        client = LangSmithClient(
            api_key=proj_config.langsmith.api_key,
            project_id=proj_config.langsmith.project_id,
            org_id=proj_config.langsmith.org_id or ""
        )

        # Fetch trace
        trace = client.get_trace_by_id(trace_id)
        if not trace:
            return [TextContent(type="text", text=f"❌ Trace '{trace_id}' non trovato.\n\nVerifica che l'ID sia corretto e che il trace esista nel progetto LangSmith.")]

        # Analizza
        analysis = client.analyze_trace(trace)

        # Ottieni child runs per errori dettagliati
        child_runs = client.get_child_runs(trace_id)

        # Formatta output
        output = []
        output.append("## 🔍 Trace Analysis\n")
        output.append(f"**URL**: [{trace_id[:8]}...]({analysis['trace_url']})")
        output.append(f"**Status**: {analysis['status']}")

        # Duration formattata
        duration_ms = analysis['duration_ms']
        if duration_ms >= 1000:
            output.append(f"**Duration**: {duration_ms/1000:.2f}s")
        else:
            output.append(f"**Duration**: {duration_ms}ms")

        if analysis.get('model'):
            output.append(f"**Model**: {analysis['model']}")

        output.append("")

        # Input
        output.append("### 📝 Input")
        input_text = trace.input[:1500] if trace.input else "(vuoto)"
        if len(trace.input) > 1500:
            input_text += "..."
        output.append(f"```\n{input_text}\n```")
        output.append("")

        # Output
        output.append("### 💬 Output")
        output_text = trace.output[:2000] if trace.output else "(vuoto)"
        if len(trace.output) > 2000:
            output_text += "..."
        output.append(f"```\n{output_text}\n```")
        output.append("")

        # Tools
        if analysis['tool_summary']['tools_used']:
            output.append(f"### 🔧 Tools ({analysis['tool_summary']['total_calls']} calls)")
            for tool_detail in analysis['tool_details'][:15]:
                icon = "❌" if tool_detail['has_error'] else "✓"
                duration_str = f"{tool_detail['duration_ms']}ms" if tool_detail['duration_ms'] < 1000 else f"{tool_detail['duration_ms']/1000:.1f}s"
                output.append(f"- {icon} **{tool_detail['name']}** ({duration_str})")

            if analysis['tool_summary']['failed_calls'] > 0:
                output.append(f"\n⚠️ **{analysis['tool_summary']['failed_calls']} tool calls fallite**")
            output.append("")

        # Tokens
        if analysis.get('tokens_used', 0) > 0:
            output.append(f"### 📊 Tokens: {analysis['tokens_used']:,}")
            output.append("")

        # Errors
        errors = [r for r in child_runs if r.get('error')]
        if errors:
            output.append("### ❌ Errori")
            for err in errors[:5]:
                err_name = err.get('name', 'unknown')
                err_msg = str(err.get('error', ''))[:300]
                output.append(f"- **{err_name}**: {err_msg}")
            output.append("")

        # Suggerimenti per debug
        if analysis['status'] != 'success' or errors:
            output.append("### 💡 Suggerimenti")
            if errors:
                output.append("- Verifica i tool che hanno fallito e i loro input")
            if analysis['duration_ms'] > 30000:
                output.append("- Il trace è lento (>30s) - considera ottimizzazioni")
            output.append("- Apri il link LangSmith per dettagli completi")

        return [TextContent(type="text", text="\n".join(output))]

    except Exception as e:
        logger.error(f"Error in debug_trace: {e}")
        return [TextContent(type="text", text=f"❌ Errore durante l'analisi del trace: {str(e)}")]
